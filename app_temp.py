# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import pandas as pd
import numpy as np
import json
import os
from datetime import datetime
import secrets
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import pytz
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from report_generator import generate_role_based_report

# ׳”׳’׳“׳¨׳× ׳׳–׳•׳¨ ׳–׳׳ ׳™׳©׳¨׳׳
IST = pytz.timezone('Asia/Jerusalem')

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # ׳׳₪׳×׳— ׳¡׳•׳“׳™ ׳׳©׳™׳׳•׳© ׳‘-session ׳•-flash

# ׳₪׳™׳׳˜׳¨ ׳×׳׳¨׳™׳›׳™׳ ׳¢׳‘׳¨׳™׳™׳
@app.template_filter('format_date')
def format_date(date):
    if isinstance(date, datetime):
        return date.strftime('%d/%m/%Y %H:%M')
    return date

# ׳§׳‘׳¦׳™ ׳ ׳×׳•׳ ׳™׳
EXCEL_FILE = 'amj.xlsx'
PAYMENT_HISTORY_FILE = 'payment_history.json'
USERS_FILE = 'users.json'

# ׳˜׳¢׳™׳ ׳× ׳§׳•׳‘׳¥ ׳׳©׳×׳׳©׳™׳ ׳׳ ׳§׳™׳™׳
def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        # ׳™׳¦׳™׳¨׳× ׳׳©׳×׳׳© ׳׳ ׳”׳ ׳¨׳׳©׳•׳ ׳™
        admin_user = {
            "username": "admin",
            "password": generate_password_hash("admin123"),
            "role": "admin",
            "name": "׳׳ ׳”׳ ׳׳¢׳¨׳›׳×",
            "entity_id": None
        }
        users = {"users": [admin_user]}
        save_users(users)
        return users

def save_users(users):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, ensure_ascii=False, indent=4)

# ׳˜׳¢׳™׳ ׳× ׳”׳™׳¡׳˜׳•׳¨׳™׳™׳× ׳×׳©׳׳•׳׳™׳
def load_payment_history():
    if os.path.exists(PAYMENT_HISTORY_FILE):
        with open(PAYMENT_HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {
            "payments": [],
            "transfers": []
        }

def save_payment_history(history):
    with open(PAYMENT_HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=4)

# ׳¨׳™׳©׳•׳ ׳×׳©׳׳•׳ ׳—׳“׳©
def record_payment(player_id, player_name, agent_name, super_agent_name, amount, payment_date=None, method="", notes="", recorded_by=""):
    history = load_payment_history()
    
    if payment_date is None:
        payment_date = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    
    payment = {
        "player_id": player_id,
        "player_name": player_name,
        "agent_name": agent_name,
        "super_agent_name": super_agent_name,
        "amount": amount,
        "payment_date": payment_date,
        "method": method,
        "notes": notes,
        "recorded_at": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
        "recorded_by": recorded_by
    }
    
    history["payments"].append(payment)
    save_payment_history(history)
    return payment

# ׳¨׳™׳©׳•׳ ׳”׳¢׳‘׳¨׳× ׳›׳¡׳₪׳™׳
def record_transfer(from_entity, from_type, to_entity, to_type, amount, transfer_date=None, notes="", recorded_by=""):
    history = load_payment_history()
    
    if transfer_date is None:
        transfer_date = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    
    transfer = {
        "from_entity": from_entity,
        "from_type": from_type,
        "to_entity": to_entity,
        "to_type": to_type,
        "amount": amount,
        "transfer_date": transfer_date,
        "notes": notes,
        "recorded_at": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
        "recorded_by": recorded_by
    }
    
    history["transfers"].append(transfer)
    save_payment_history(history)
    return transfer

# ׳˜׳¢׳™׳ ׳× ׳ ׳×׳•׳ ׳™׳ ׳׳”׳׳§׳¡׳
def load_excel_data():
    try:
        # ׳§׳¨׳™׳׳× ׳’׳™׳׳™׳•׳ game stats
        game_stats_df = pd.read_excel(EXCEL_FILE, sheet_name='game stats')
        
        # ׳§׳¨׳™׳׳× ׳›׳ ׳”׳’׳™׳׳™׳•׳ ׳•׳×
        xl = pd.ExcelFile(EXCEL_FILE)
        sheets = xl.sheet_names
        
        # ׳׳™׳₪׳•׳™ ׳©׳—׳§׳ ׳™׳, ׳׳™׳™׳’'׳ ׳˜׳™׳ ׳•׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜׳™׳
        players = game_stats_df[['׳§׳•׳“ ׳©׳—׳§׳', '׳©׳ ׳©׳—׳§׳']].drop_duplicates().dropna().to_dict('records')
        agents = game_stats_df['׳©׳ ׳׳™׳™׳’׳ ׳˜'].dropna().unique().tolist()
        super_agents = game_stats_df['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'].dropna().unique().tolist()
        
        return {
            'game_stats': game_stats_df,
            'sheets': sheets,
            'players': players,
            'agents': agents,
            'super_agents': super_agents
        }
    except Exception as e:
        print(f"׳©׳’׳™׳׳” ׳‘׳˜׳¢׳™׳ ׳× ׳”׳ ׳×׳•׳ ׳™׳: {str(e)}")
        return None

# ׳—׳™׳©׳•׳‘ ׳¡׳™׳›׳•׳׳™׳ ׳׳“׳©׳‘׳•׳¨׳“
def calculate_dashboard_data():
    try:
        # ׳˜׳¢׳™׳ ׳× ׳”׳™׳¡׳˜׳•׳¨׳™׳™׳× ׳×׳©׳׳•׳׳™׳
        history = load_payment_history()
        payments = history.get('payments', [])
        transfers = history.get('transfers', [])
        
        # ׳˜׳¢׳™׳ ׳× ׳ ׳×׳•׳ ׳™ Excel
        excel_data = load_excel_data()
        if not excel_data:
            return {}
            
        # ׳—׳™׳©׳•׳‘ ׳¡׳”"׳› ׳׳’׳‘׳™׳™׳”
        try:
            # ׳ ׳™׳¡׳™׳•׳ ׳׳˜׳¢׳•׳ ׳’׳™׳׳™׳•׳ ׳¡׳™׳›׳•׳׳™ ׳’׳‘׳™׳™׳”
            collection_summary = pd.read_excel(EXCEL_FILE, sheet_name='׳¡׳™׳›׳•׳׳™ ׳’׳‘׳™׳”')
            total_to_collect = collection_summary['׳¡׳”"׳› ׳׳’׳‘׳™׳”'].sum()
        except:
            # ׳׳ ׳׳™׳ ׳’׳™׳׳™׳•׳ ׳›׳–׳”, ׳׳—׳©׳‘ ׳׳”׳’׳™׳׳™׳•׳ ׳”׳¨׳׳©׳™
            game_stats = excel_data['game_stats']
            total_to_collect = game_stats['׳‘׳׳׳ ׳¡'].sum()
        
        # ׳—׳™׳©׳•׳‘ ׳¡׳”"׳› ׳©׳©׳•׳׳
        total_paid = sum(payment['amount'] for payment in payments)
        
        # ׳—׳™׳©׳•׳‘ ׳׳¡׳₪׳¨ ׳©׳—׳§׳ ׳™׳
        players_count = len(excel_data['players'])
        
        # ׳—׳™׳©׳•׳‘ ׳׳¡׳₪׳¨ ׳׳™׳™׳’'׳ ׳˜׳™׳
        agents_count = len(excel_data['agents'])
        
        # ׳—׳™׳©׳•׳‘ ׳¡׳”"׳› ׳”׳¢׳‘׳¨׳•׳×
        total_transfers = sum(transfer['amount'] for transfer in transfers)
        
        # ׳×׳©׳׳•׳׳™׳ ׳׳—׳¨׳•׳ ׳™׳
        last_payments = sorted(payments, key=lambda x: x['recorded_at'], reverse=True)[:5]
        
        # ׳”׳¢׳‘׳¨׳•׳× ׳׳—׳¨׳•׳ ׳•׳×
        last_transfers = sorted(transfers, key=lambda x: x['recorded_at'], reverse=True)[:5]
        
        return {
            'total_to_collect': total_to_collect,
            'total_paid': total_paid,
            'players_count': players_count,
            'agents_count': agents_count,
            'total_transfers': total_transfers,
            'last_payments': last_payments,
            'last_transfers': last_transfers
        }
    except Exception as e:
        print(f"׳©׳’׳™׳׳” ׳‘׳—׳™׳©׳•׳‘ ׳ ׳×׳•׳ ׳™ ׳“׳©׳‘׳•׳¨׳“: {str(e)}")
        return {}

# ׳¢׳“׳›׳•׳ ׳§׳•׳‘׳¥ ׳”׳׳§׳¡׳ ׳¢׳ ׳ ׳×׳•׳ ׳™ ׳×׳©׳׳•׳׳™׳
def update_excel_with_payments(output_file=None):
    try:
        if output_file is None:
            # ׳™׳¦׳™׳¨׳× ׳§׳•׳‘׳¥ ׳₪׳׳˜ ׳—׳“׳© ׳¢׳ ׳‘׳¡׳™׳¡ ׳§׳•׳‘׳¥ ׳”׳׳§׳•׳¨
            filename, ext = os.path.splitext(EXCEL_FILE)
            output_file = f"{filename}_updated{ext}"
        
        # ׳”׳¢׳×׳§׳× ׳”׳׳§׳•׳¨ ׳׳§׳•׳‘׳¥ ׳”׳—׳“׳©
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # ׳˜׳¢׳™׳ ׳× ׳”׳™׳¡׳˜׳•׳¨׳™׳™׳× ׳×׳©׳׳•׳׳™׳
        history = load_payment_history()
        payments = history.get('payments', [])
        
        # ׳¢׳“׳›׳•׳ ׳¢׳׳•׳“׳× '׳©׳•׳׳' ׳‘׳›׳ ׳”׳’׳™׳׳™׳•׳ ׳•׳×
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name == "Sheet" or not any(cell.value == "׳©׳•׳׳" for row in ws.iter_rows(max_row=3) for cell in row):
                continue
                
            # ׳׳¦׳™׳׳× ׳”׳¢׳׳•׳“׳•׳× ׳”׳¨׳׳•׳•׳ ׳˜׳™׳•׳×
            header_row = 2  # ׳©׳•׳¨׳× ׳”׳›׳•׳×׳¨׳•׳×
            paid_col = None
            player_id_col = None
            player_name_col = None
            
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=header_row, column=col).value == "׳©׳•׳׳":
                    paid_col = col
                elif ws.cell(row=header_row, column=col).value == "׳§׳•׳“ ׳©׳—׳§׳":
                    player_id_col = col
                elif ws.cell(row=header_row, column=col).value == "׳©׳ ׳©׳—׳§׳":
                    player_name_col = col
            
            if not (paid_col and player_id_col and player_name_col):
                continue
                
            # ׳¢׳“׳›׳•׳ ׳©׳•׳¨׳•׳× ׳”׳©׳—׳§׳ ׳™׳
            for row in range(3, ws.max_row + 1):
                player_id = ws.cell(row=row, column=player_id_col).value
                player_name = ws.cell(row=row, column=player_name_col).value
                
                if player_id and player_name:
                    # ׳—׳™׳©׳•׳‘ ׳¡׳”"׳› ׳©׳©׳•׳׳ ׳׳₪׳™ ׳©׳—׳§׳
                    total_paid = sum(
                        payment['amount'] for payment in payments 
                        if str(payment['player_id']) == str(player_id) and payment['player_name'] == player_name
                    )
                    
                    # ׳¢׳“׳›׳•׳ ׳”׳×׳
                    ws.cell(row=row, column=paid_col).value = total_paid
        
        # ׳©׳׳™׳¨׳× ׳”׳§׳•׳‘׳¥
        wb.save(output_file)
        return output_file
    except Exception as e:
        print(f"׳©׳’׳™׳׳” ׳‘׳¢׳“׳›׳•׳ ׳”׳§׳•׳‘׳¥: {str(e)}")
        return None

# ׳₪׳•׳ ׳§׳¦׳™׳™׳× ׳׳™׳׳•׳× ׳׳©׳×׳׳© ׳¢׳‘׳•׳¨ ׳¦׳“ ׳©׳¨׳×
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('׳ ׳ ׳׳”׳×׳—׳‘׳¨ ׳›׳“׳™ ׳׳’׳©׳× ׳׳“׳£ ׳–׳”', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# ׳₪׳•׳ ׳§׳¦׳™׳™׳× ׳׳™׳׳•׳× ׳”׳¨׳©׳׳× ׳׳ ׳”׳
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or session['user']['role'] != 'admin':
            flash('׳׳™׳ ׳׳ ׳”׳¨׳©׳׳” ׳׳’׳©׳× ׳׳“׳£ ׳–׳”', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# ׳₪׳•׳ ׳§׳¦׳™׳™׳× ׳׳™׳׳•׳× ׳”׳¨׳©׳׳× ׳׳™׳™׳’'׳ ׳˜ ׳׳• ׳׳ ׳”׳
def agent_or_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or (session['user']['role'] != 'admin' and session['user']['role'] != 'agent' and session['user']['role'] != 'super_agent'):
            flash('׳׳™׳ ׳׳ ׳”׳¨׳©׳׳” ׳׳’׳©׳× ׳׳“׳£ ׳–׳”', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

#
# ׳ ׳™׳×׳•׳‘׳™ ׳”׳׳₪׳׳™׳§׳¦׳™׳” (Routes)
#

# ׳“׳£ ׳”׳‘׳™׳× - ׳׳¢׳‘׳¨ ׳׳“׳©׳‘׳•׳¨׳“ ׳׳ ׳”׳׳©׳×׳׳© ׳׳—׳•׳‘׳¨
@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# ׳“׳£ ׳”׳×׳—׳‘׳¨׳•׳×
@app.route('/login', methods=['GET', 'POST'])
def login():
    # ׳׳ ׳”׳׳©׳×׳׳© ׳›׳‘׳¨ ׳׳—׳•׳‘׳¨, ׳”׳₪׳ ׳” ׳׳“׳©׳‘׳•׳¨׳“
    if 'user' in session:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # ׳׳™׳׳•׳× ׳׳©׳×׳׳©
        users = load_users()
        for user in users['users']:
            if user['username'] == username and check_password_hash(user['password'], password):
                # ׳©׳׳™׳¨׳× ׳₪׳¨׳˜׳™ ׳”׳׳©׳×׳׳© ׳‘-session
                session['user'] = {
                    'username': user['username'],
                    'name': user['name'],
                    'role': user['role'],
                    'entity_id': user['entity_id']
                }
                
                flash(f'׳‘׳¨׳•׳ ׳”׳‘׳, {user["name"]}!', 'success')
                
                # ׳”׳₪׳ ׳™׳” ׳׳¢׳׳•׳“ ׳”׳‘׳ ׳׳• ׳׳“׳©׳‘׳•׳¨׳“
                next_page = request.args.get('next')
                if next_page:
                    return redirect(next_page)
                return redirect(url_for('dashboard'))
        
        flash('׳©׳ ׳׳©׳×׳׳© ׳׳• ׳¡׳™׳¡׳׳” ׳©׳’׳•׳™׳™׳', 'danger')
    
    return render_template('login.html')

# ׳™׳¦׳™׳׳” ׳׳”׳׳¢׳¨׳›׳×
@app.route('/logout')
def logout():
    session.pop('user', None)
    flash('׳”׳×׳ ׳×׳§׳× ׳׳”׳׳¢׳¨׳›׳× ׳‘׳”׳¦׳׳—׳”', 'success')
    return redirect(url_for('login'))

# ׳׳•׳— ׳׳—׳•׳•׳ ׳™׳ (׳“׳©׳‘׳•׳¨׳“)
@app.route('/dashboard')
@login_required
def dashboard():
    dashboard_data = calculate_dashboard_data()
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    return render_template('dashboard.html', 
                          dashboard_data=dashboard_data, 
                          user_role=user_role,
                          user_entity_id=user_entity_id)

# ׳¨׳©׳™׳׳× ׳©׳—׳§׳ ׳™׳
@app.route('/players')
@login_required
def players():
    excel_data = load_excel_data()
    players_list = excel_data['players']
    
    # ׳”׳×׳׳׳” ׳׳₪׳™ ׳”׳¨׳©׳׳•׳×
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'agent':
        # ׳¡׳™׳ ׳•׳ ׳©׳—׳§׳ ׳™׳ ׳”׳©׳™׳™׳›׳™׳ ׳׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        game_stats = excel_data['game_stats']
        agent_players = game_stats[game_stats['׳©׳ ׳׳™׳™׳’׳ ׳˜'] == user_entity_id]['׳§׳•׳“ ׳©׳—׳§׳'].unique()
        players_list = [p for p in players_list if p['׳§׳•׳“ ׳©׳—׳§׳'] in agent_players]
    elif user_role == 'super_agent':
        # ׳¡׳™׳ ׳•׳ ׳©׳—׳§׳ ׳™׳ ׳”׳©׳™׳™׳›׳™׳ ׳׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        game_stats = excel_data['game_stats']
        super_agent_players = game_stats[game_stats['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'] == user_entity_id]['׳§׳•׳“ ׳©׳—׳§׳'].unique()
        players_list = [p for p in players_list if p['׳§׳•׳“ ׳©׳—׳§׳'] in super_agent_players]
    
    # ׳˜׳¢׳™׳ ׳× ׳ ׳×׳•׳ ׳™ ׳×׳©׳׳•׳׳™׳
    history = load_payment_history()
    payments = history.get('payments', [])
    
    # ׳—׳™׳©׳•׳‘ ׳¡׳”"׳› ׳©׳©׳•׳׳ ׳׳›׳ ׳©׳—׳§׳
    for player in players_list:
        player_id = player['׳§׳•׳“ ׳©׳—׳§׳']
        player['total_paid'] = sum(
            payment['amount'] for payment in payments 
            if str(payment['player_id']) == str(player_id)
        )
    
    return render_template('players.html', 
                          players=players_list, 
                          user_role=user_role)

# ׳₪׳¨׳˜׳™ ׳©׳—׳§׳
@app.route('/player/<player_id>')
@login_required
def player_details(player_id):
    excel_data = load_excel_data()
    game_stats = excel_data['game_stats']
    
    # ׳׳¦׳™׳׳× ׳ ׳×׳•׳ ׳™ ׳”׳©׳—׳§׳
    player_data = game_stats[game_stats['׳§׳•׳“ ׳©׳—׳§׳'].astype(str) == str(player_id)].to_dict('records')
    
    if not player_data:
        flash('׳©׳—׳§׳ ׳׳ ׳ ׳׳¦׳', 'warning')
        return redirect(url_for('players'))
    
    # ׳׳™׳“׳¢ ׳‘׳¡׳™׳¡׳™ ׳¢׳ ׳”׳©׳—׳§׳
    player_info = {
        'id': player_id,
        'name': player_data[0]['׳©׳ ׳©׳—׳§׳'],
        'agent': player_data[0]['׳©׳ ׳׳™׳™׳’׳ ׳˜'],
        'super_agent': player_data[0]['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜']
    }
    
    # ׳˜׳¢׳™׳ ׳× ׳”׳™׳¡׳˜׳•׳¨׳™׳™׳× ׳×׳©׳׳•׳׳™׳
    history = load_payment_history()
    
    # ׳¡׳™׳ ׳•׳ ׳×׳©׳׳•׳׳™׳ ׳©׳ ׳”׳©׳—׳§׳
    player_payments = [
        payment for payment in history.get('payments', [])
        if str(payment['player_id']) == str(player_id)
    ]
    
    # ׳—׳™׳©׳•׳‘ ׳¡׳”"׳› ׳©׳©׳•׳׳
    total_paid = sum(payment['amount'] for payment in player_payments)
    
    return render_template('player_details.html', 
                          player=player_info, 
                          payments=player_payments,
                          total_paid=total_paid)

# ׳¨׳©׳™׳׳× ׳׳™׳™׳’'׳ ׳˜׳™׳
@app.route('/agents')
@login_required
def agents():
    excel_data = load_excel_data()
    agents_list = excel_data['agents']
    
    # ׳”׳×׳׳׳” ׳׳₪׳™ ׳”׳¨׳©׳׳•׳×
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'super_agent':
        # ׳¡׳™׳ ׳•׳ ׳׳™׳™׳’'׳ ׳˜׳™׳ ׳”׳©׳™׳™׳›׳™׳ ׳׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        game_stats = excel_data['game_stats']
        super_agent_agents = game_stats[game_stats['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'] == user_entity_id]['׳©׳ ׳׳™׳™׳’׳ ׳˜'].unique()
        agents_list = [a for a in agents_list if a in super_agent_agents]
    
    # ׳™׳¦׳™׳¨׳× ׳¨׳©׳™׳׳× ׳׳•׳‘׳™׳™׳§׳˜׳™׳ ׳¢׳ ׳׳™׳“׳¢ ׳ ׳•׳¡׳£ ׳׳›׳ ׳׳™׳™׳’'׳ ׳˜
    agents_data = []
    game_stats = excel_data['game_stats']
    
    for agent in agents_list:
        agent_players = game_stats[game_stats['׳©׳ ׳׳™׳™׳’׳ ׳˜'] == agent]['׳§׳•׳“ ׳©׳—׳§׳'].unique()
        players_count = len(agent_players)
        
        # ׳׳¦׳™׳׳× ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜
        super_agent = game_stats[game_stats['׳©׳ ׳׳™׳™׳’׳ ׳˜'] == agent]['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'].dropna().unique()
        super_agent_name = super_agent[0] if len(super_agent) > 0 else ""
        
        agents_data.append({
            'name': agent,
            'super_agent': super_agent_name,
            'players_count': players_count
        })
    
    return render_template('agents.html', 
                          agents=agents_data, 
                          user_role=user_role)

# ׳₪׳¨׳˜׳™ ׳׳™׳™׳’'׳ ׳˜
@app.route('/agent/<agent_name>')
@login_required
def agent_details(agent_name):
    excel_data = load_excel_data()
    game_stats = excel_data['game_stats']
    
    # ׳׳¦׳™׳׳× ׳©׳—׳§׳ ׳™׳ ׳©׳ ׳”׳׳™׳™׳’'׳ ׳˜
    agent_data = game_stats[game_stats['׳©׳ ׳׳™׳™׳’׳ ׳˜'] == agent_name]
    
    if len(agent_data) == 0:
        flash('׳׳™׳™׳’\'׳ ׳˜ ׳׳ ׳ ׳׳¦׳', 'warning')
        return redirect(url_for('agents'))
    
    # ׳¨׳©׳™׳׳× ׳©׳—׳§׳ ׳™׳
    players = agent_data[['׳§׳•׳“ ׳©׳—׳§׳', '׳©׳ ׳©׳—׳§׳']].drop_duplicates().to_dict('records')
    
    # ׳׳™׳“׳¢ ׳¢׳ ׳”׳׳™׳™׳’'׳ ׳˜
    super_agent = agent_data['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'].dropna().unique()
    agent_info = {
        'name': agent_name,
        'super_agent': super_agent[0] if len(super_agent) > 0 else "",
        'players_count': len(players)
    }
    
    # ׳˜׳¢׳™׳ ׳× ׳”׳™׳¡׳˜׳•׳¨׳™׳™׳× ׳×׳©׳׳•׳׳™׳
    history = load_payment_history()
    
    # ׳¡׳™׳ ׳•׳ ׳×׳©׳׳•׳׳™׳ ׳©׳ ׳”׳׳™׳™׳’'׳ ׳˜
    agent_payments = [
        payment for payment in history.get('payments', [])
        if payment['agent_name'] == agent_name
    ]
    
    # ׳—׳™׳©׳•׳‘ ׳¡׳”"׳› ׳©׳©׳•׳׳
    total_paid = sum(payment['amount'] for payment in agent_payments)
    
    return render_template('agent_details.html', 
                          agent=agent_info,
                          players=players,
                          payments=agent_payments,
                          total_paid=total_paid)

# ׳¨׳©׳™׳׳× ׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜׳™׳
@app.route('/super-agents')
@login_required
def super_agents():
    excel_data = load_excel_data()
    super_agents_list = excel_data['super_agents']
    
    # ׳”׳×׳׳׳” ׳׳₪׳™ ׳”׳¨׳©׳׳•׳×
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'super_agent':
        # ׳”׳¦׳’ ׳¨׳§ ׳׳× ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        super_agents_list = [sa for sa in super_agents_list if sa == user_entity_id]
    
    # ׳™׳¦׳™׳¨׳× ׳¨׳©׳™׳׳× ׳׳•׳‘׳™׳™׳§׳˜׳™׳ ׳¢׳ ׳׳™׳“׳¢ ׳ ׳•׳¡׳£ ׳׳›׳ ׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜
    super_agents_data = []
    game_stats = excel_data['game_stats']
    
    for super_agent in super_agents_list:
        agent_count = len(game_stats[game_stats['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'] == super_agent]['׳©׳ ׳׳™׳™׳’׳ ׳˜'].unique())
        player_count = len(game_stats[game_stats['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'] == super_agent]['׳§׳•׳“ ׳©׳—׳§׳'].unique())
        
        super_agents_data.append({
            'name': super_agent,
            'agents_count': agent_count,
            'players_count': player_count
        })
    
    return render_template('super_agents.html', 
                          super_agents=super_agents_data, 
                          user_role=user_role)

# ׳₪׳¨׳˜׳™ ׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜
@app.route('/super-agent/<super_agent_name>')
@login_required
def super_agent_details(super_agent_name):
    excel_data = load_excel_data()
    game_stats = excel_data['game_stats']
    
    # ׳׳¦׳™׳׳× ׳ ׳×׳•׳ ׳™ ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜
    super_agent_data = game_stats[game_stats['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'] == super_agent_name]
    
    if len(super_agent_data) == 0:
        flash('׳¡׳•׳₪׳¨-׳׳™׳™׳’\'׳ ׳˜ ׳׳ ׳ ׳׳¦׳', 'warning')
        return redirect(url_for('super_agents'))
    
    # ׳¨׳©׳™׳׳× ׳׳™׳™׳’'׳ ׳˜׳™׳
    agents = super_agent_data['׳©׳ ׳׳™׳™׳’׳ ׳˜'].dropna().unique().tolist()
    
    # ׳׳™׳“׳¢ ׳¢׳ ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜
    super_agent_info = {
        'name': super_agent_name,
        'agents_count': len(agents),
        'players_count': len(super_agent_data['׳§׳•׳“ ׳©׳—׳§׳'].unique())
    }
    
    # ׳˜׳¢׳™׳ ׳× ׳”׳™׳¡׳˜׳•׳¨׳™׳™׳× ׳×׳©׳׳•׳׳™׳
    history = load_payment_history()
    
    # ׳¡׳™׳ ׳•׳ ׳×׳©׳׳•׳׳™׳ ׳©׳ ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜
    super_agent_payments = [
        payment for payment in history.get('payments', [])
        if payment['super_agent_name'] == super_agent_name
    ]
    
    # ׳—׳™׳©׳•׳‘ ׳¡׳”"׳› ׳©׳©׳•׳׳
    total_paid = sum(payment['amount'] for payment in super_agent_payments)
    
    return render_template('super_agent_details.html', 
                          super_agent=super_agent_info,
                          agents=agents,
                          payments=super_agent_payments,
                          total_paid=total_paid)

# ׳¨׳™׳©׳•׳ ׳×׳©׳׳•׳ ׳—׳“׳©
@app.route('/add-payment', methods=['GET', 'POST'])
@agent_or_admin_required
def add_payment():
    excel_data = load_excel_data()
    players = excel_data['players']
    agents = excel_data['agents']
    super_agents = excel_data['super_agents']
    
    # ׳”׳×׳׳׳” ׳׳₪׳™ ׳”׳¨׳©׳׳•׳×
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'agent':
        # ׳”׳¦׳’ ׳¨׳§ ׳©׳—׳§׳ ׳™׳ ׳©׳ ׳”׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        game_stats = excel_data['game_stats']
        agent_players = game_stats[game_stats['׳©׳ ׳׳™׳™׳’׳ ׳˜'] == user_entity_id]['׳§׳•׳“ ׳©׳—׳§׳'].unique()
        players = [p for p in players if p['׳§׳•׳“ ׳©׳—׳§׳'] in agent_players]
        
        # ׳”׳’׳‘׳ ׳׳× ׳”׳׳™׳™׳’'׳ ׳˜ ׳׳”׳™׳•׳× ׳”׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        agents = [user_entity_id]
        
        # ׳”׳’׳‘׳ ׳׳× ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳׳”׳™׳•׳× ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳©׳ ׳”׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        super_agent = game_stats[game_stats['׳©׳ ׳׳™׳™׳’׳ ׳˜'] == user_entity_id]['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'].dropna().unique()
        super_agents = [super_agent[0]] if len(super_agent) > 0 else []
    elif user_role == 'super_agent':
        # ׳”׳¦׳’ ׳¨׳§ ׳©׳—׳§׳ ׳™׳ ׳©׳ ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        game_stats = excel_data['game_stats']
        super_agent_players = game_stats[game_stats['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'] == user_entity_id]['׳§׳•׳“ ׳©׳—׳§׳'].unique()
        players = [p for p in players if p['׳§׳•׳“ ׳©׳—׳§׳'] in super_agent_players]
        
        # ׳”׳’׳‘׳ ׳׳× ׳”׳׳™׳™׳’'׳ ׳˜ ׳׳”׳™׳•׳× ׳”׳׳™׳™׳’'׳ ׳˜׳™׳ ׳©׳ ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        agents = game_stats[game_stats['׳©׳ ׳¡׳•׳₪׳¨ ׳׳™׳™׳’׳ ׳˜'] == user_entity_id]['׳©׳ ׳׳™׳™׳’׳ ׳˜'].dropna().unique().tolist()
        
        # ׳”׳’׳‘׳ ׳׳× ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳׳”׳™׳•׳× ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        super_agents = [user_entity_id]
    
    if request.method == 'POST':
        player_id = request.form.get('player_id')
        player_name = request.form.get('player_name')
        agent_name = request.form.get('agent_name')
        super_agent_name = request.form.get('super_agent_name')
        amount = float(request.form.get('amount'))
        payment_date = request.form.get('payment_date')
        method = request.form.get('method')
        notes = request.form.get('notes')
        
        # ׳×׳™׳§׳•׳ ׳¢׳¨׳›׳™׳ ׳¨׳™׳§׳™׳
        if not payment_date:
            payment_date = datetime.now(IST).strftime("%Y-%m-%d")
        
        # ׳¨׳™׳©׳•׳ ׳”׳×׳©׳׳•׳
        payment = record_payment(
            player_id=player_id,
            player_name=player_name,
            agent_name=agent_name,
            super_agent_name=super_agent_name,
            amount=amount,
            payment_date=payment_date,
            method=method,
            notes=notes,
            recorded_by=session['user']['username']
        )
        
        flash('׳”׳×׳©׳׳•׳ ׳ ׳¨׳©׳ ׳‘׳”׳¦׳׳—׳”', 'success')
        return redirect(url_for('payments'))
    
    return render_template('add_payment.html', 
                          players=players, 
                          agents=agents, 
                          super_agents=super_agents,
                          user_role=user_role)

# ׳¨׳™׳©׳•׳ ׳”׳¢׳‘׳¨׳× ׳›׳¡׳₪׳™׳
@app.route('/add-transfer', methods=['GET', 'POST'])
@agent_or_admin_required
def add_transfer():
    excel_data = load_excel_data()
    agents = excel_data['agents']
    super_agents = excel_data['super_agents']
    
    # ׳”׳×׳׳׳” ׳׳₪׳™ ׳”׳¨׳©׳׳•׳×
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role != 'admin':
        # ׳¨׳§ ׳׳ ׳”׳ ׳™׳›׳•׳ ׳׳¨׳©׳•׳ ׳”׳¢׳‘׳¨׳•׳×
        flash('׳׳™׳ ׳׳ ׳”׳¨׳©׳׳” ׳׳‘׳¦׳¢ ׳₪׳¢׳•׳׳” ׳–׳•', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        from_type = request.form.get('from_type')
        from_entity = request.form.get('from_entity')
        to_type = request.form.get('to_type')
        to_entity = request.form.get('to_entity')
        amount = float(request.form.get('amount'))
        transfer_date = request.form.get('transfer_date')
        notes = request.form.get('notes')
        
        # ׳×׳™׳§׳•׳ ׳¢׳¨׳›׳™׳ ׳¨׳™׳§׳™׳
        if not transfer_date:
            transfer_date = datetime.now(IST).strftime("%Y-%m-%d")
        
        # ׳¨׳™׳©׳•׳ ׳”׳”׳¢׳‘׳¨׳”
        transfer = record_transfer(
            from_entity=from_entity,
            from_type=from_type,
            to_entity=to_entity,
            to_type=to_type,
            amount=amount,
            transfer_date=transfer_date,
            notes=notes,
            recorded_by=session['user']['username']
        )
        
        flash('׳”׳¢׳‘׳¨׳× ׳”׳›׳¡׳₪׳™׳ ׳ ׳¨׳©׳׳” ׳‘׳”׳¦׳׳—׳”', 'success')
        return redirect(url_for('transfers'))
    
    return render_template('add_transfer.html', 
                          agents=agents, 
                          super_agents=super_agents)

# ׳¨׳©׳™׳׳× ׳×׳©׳׳•׳׳™׳
@app.route('/payments')
@login_required
def payments():
    # ׳˜׳¢׳™׳ ׳× ׳”׳™׳¡׳˜׳•׳¨׳™׳™׳× ׳×׳©׳׳•׳׳™׳
    history = load_payment_history()
    payments_list = history.get('payments', [])
    
    # ׳”׳×׳׳׳” ׳׳₪׳™ ׳”׳¨׳©׳׳•׳×
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'agent':
        # ׳¡׳™׳ ׳•׳ ׳×׳©׳׳•׳׳™׳ ׳©׳ ׳”׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        payments_list = [p for p in payments_list if p['agent_name'] == user_entity_id]
    elif user_role == 'super_agent':
        # ׳¡׳™׳ ׳•׳ ׳×׳©׳׳•׳׳™׳ ׳©׳ ׳”׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳”׳ ׳•׳›׳—׳™
        payments_list = [p for p in payments_list if p['super_agent_name'] == user_entity_id]
    
    # ׳׳™׳•׳ ׳׳₪׳™ ׳×׳׳¨׳™׳ ׳¨׳™׳©׳•׳ (׳׳”׳—׳“׳© ׳׳™׳©׳)
    payments_list = sorted(payments_list, key=lambda x: x['recorded_at'], reverse=True)
    
    return render_template('payments.html', 
                          payments=payments_list,
                          user_role=user_role)

# ׳¨׳©׳™׳׳× ׳”׳¢׳‘׳¨׳•׳×
@app.route('/transfers')
@login_required
def transfers():
    # ׳˜׳¢׳™׳ ׳× ׳”׳™׳¡׳˜׳•׳¨׳™׳™׳× ׳”׳¢׳‘׳¨׳•׳×
    history = load_payment_history()
    transfers_list = history.get('transfers', [])
    
    # ׳׳™׳•׳ ׳׳₪׳™ ׳×׳׳¨׳™׳ ׳¨׳™׳©׳•׳ (׳׳”׳—׳“׳© ׳׳™׳©׳)
    transfers_list = sorted(transfers_list, key=lambda x: x['recorded_at'], reverse=True)
    
    # ׳”׳×׳׳׳” ׳׳₪׳™ ׳”׳¨׳©׳׳•׳×
    user_role = session['user']['role']
    
    if user_role != 'admin':
        # ׳¨׳§ ׳׳ ׳”׳ ׳¨׳•׳׳” ׳׳× ׳›׳ ׳”׳”׳¢׳‘׳¨׳•׳×
        transfers_list = []
    
    return render_template('transfers.html', 
                          transfers=transfers_list,
                          user_role=user_role)

# ׳“׳£ ׳“׳•׳—׳•׳×
@login_required
def reports():
    excel_data = load_excel_data()
    sheets = excel_data['sheets']
    
    # ׳”׳¢׳‘׳¨׳× ׳׳™׳“׳¢ ׳¢׳ ׳׳©׳×׳׳©
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    return render_template('reports.html', 
                          sheets=sheets,
                          user_role=user_role,
                          user_entity_id=user_entity_id)

# ׳×׳¦׳•׳’׳× ׳“׳•׳— ׳¡׳₪׳¦׳™׳₪׳™
@login_required
def view_report(sheet_name):
    try:
        # ׳§׳¨׳™׳׳× ׳”׳’׳™׳׳™׳•׳ ׳”׳¡׳₪׳¦׳™׳₪׳™
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        
        # ׳”׳׳¨׳” ׳׳˜׳‘׳׳× HTML
        table = df.to_html(classes='table table-striped table-bordered table-hover', index=False)
        
        return render_template('view_report.html', 
                              sheet_name=sheet_name,
                              table=table,
                              user_role=session['user']['role'])
    except Exception as e:
        flash(f'׳©׳’׳™׳׳” ׳‘׳˜׳¢׳™׳ ׳× ׳”׳“׳•׳—: {str(e)}', 'danger')
        return redirect(url_for('reports'))

# ׳™׳¦׳™׳¨׳× ׳“׳•׳— ׳׳¢׳•׳“׳›׳
@app.route('/generate_report')
@login_required
def generate_report():
    try:
        # ׳™׳¦׳™׳¨׳× ׳“׳•׳— ׳׳₪׳™ ׳×׳₪׳§׳™׳“ ׳”׳׳©׳×׳׳©
        user_role = session['user']['role']
        user_entity_id = session['user']['entity_id']
        
        # ׳™׳¦׳™׳¨׳× ׳”׳“׳•׳—
        output_file = generate_role_based_report(user_role, user_entity_id)
        
        if output_file and os.path.exists(output_file):
            # ׳”׳—׳–׳¨׳× ׳”׳§׳•׳‘׳¥ ׳׳”׳•׳¨׳“׳”
            return send_file(
                output_file,
                as_attachment=True,
                download_name=f"{user_role}_report.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('׳©׳’׳™׳׳” ׳‘׳™׳¦׳™׳¨׳× ׳”׳“׳•׳—', 'danger')
            return redirect(url_for('reports'))
    
    except Exception as e:
        flash(f'׳©׳’׳™׳׳” ׳‘׳™׳¦׳™׳¨׳× ׳”׳“׳•׳—: {str(e)}', 'danger')
        return redirect(url_for('reports'))

# ׳“׳£ ׳“׳•׳—׳•׳×
@app.route('/reports')
@login_required
def reports():
    # ׳–׳™׳”׳•׳™ ׳”׳“׳•׳—׳•׳× ׳”׳§׳™׳™׳׳™׳
    report_files = []
    for file in os.listdir('.'):
        if file.endswith('.xlsx') and file not in [EXCEL_FILE, 'amj_updated.xlsx']:
            report_files.append({
                'filename': file,
                'display_name': file.replace('_', ' ').replace('.xlsx', ''),
                'category': '׳“׳•׳— ׳׳•׳×׳׳ ׳׳™׳©׳™׳×',
                'modified': datetime.fromtimestamp(os.path.getmtime(file))
            })
    
    # ׳׳™׳•׳ ׳”׳“׳•׳—׳•׳× ׳׳₪׳™ ׳×׳׳¨׳™׳ ׳¢׳“׳›׳•׳ (׳”׳—׳“׳© ׳‘׳™׳•׳×׳¨ ׳§׳•׳“׳)
    report_files.sort(key=lambda x: x['modified'], reverse=True)
    
    # ׳–׳™׳”׳•׳™ ׳§׳˜׳’׳•׳¨׳™׳•׳× ׳“׳•׳—׳•׳×
    admin_reports = [r for r in report_files if 'admin' in r['filename'].lower()]
    super_agent_reports = [r for r in report_files if 'super_agent' in r['filename'].lower()]
    agent_reports = [r for r in report_files if 'agent_' in r['filename'].lower() and 'super' not in r['filename'].lower()]
    other_reports = [r for r in report_files if not any(x in r['filename'].lower() for x in ['admin', 'super_agent', 'agent_'])]
    
    # ׳”׳¢׳‘׳¨׳× ׳׳™׳“׳¢ ׳¢׳ ׳׳©׳×׳׳©
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id'] if 'entity_id' in session['user'] else None
    
    # ׳¡׳™׳ ׳•׳ ׳“׳•׳—׳•׳× ׳׳₪׳™ ׳×׳₪׳§׳™׳“ ׳”׳׳©׳×׳׳©
    if user_role == 'admin':
        # ׳׳ ׳”׳ ׳¨׳•׳׳” ׳׳× ׳›׳ ׳”׳“׳•׳—׳•׳×
        pass
    elif user_role == 'super_agent':
        # ׳¡׳•׳₪׳¨-׳׳™׳™׳’'׳ ׳˜ ׳¨׳•׳׳” ׳¨׳§ ׳׳× ׳”׳“׳•׳—׳•׳× ׳©׳׳•
        super_agent_reports = [r for r in super_agent_reports 
                              if user_entity_id and user_entity_id.lower() in r['filename'].lower()]
        # ׳•׳׳™׳™׳’'׳ ׳˜׳™׳ ׳”׳׳©׳•׳™׳›׳™׳ ׳׳׳™׳• - ׳›׳׳ ׳ ׳“׳¨׳© ׳§׳•׳“ ׳ ׳•׳¡׳£ ׳׳׳¦׳•׳ ׳׳× ׳”׳׳™׳™׳’'׳ ׳˜׳™׳ ׳”׳׳©׳•׳™׳›׳™׳
        admin_reports = []
    elif user_role == 'agent':
        # ׳׳™׳™׳’'׳ ׳˜ ׳¨׳•׳׳” ׳¨׳§ ׳׳× ׳”׳“׳•׳—׳•׳× ׳©׳׳•
        agent_reports = [r for r in agent_reports 
                        if user_entity_id and user_entity_id.lower() in r['filename'].lower()]
        super_agent_reports = []
        admin_reports = []
    
    return render_template('reports.html', 
                          admin_reports=admin_reports,
                          super_agent_reports=super_agent_reports,
                          agent_reports=agent_reports,
                          other_reports=other_reports,
                          user_role=user_role,
                          user_entity_id=user_entity_id)

# ׳×׳¦׳•׳’׳× ׳“׳•׳— ׳¡׳₪׳¦׳™׳₪׳™
@app.route('/view_report/<filename>')
@login_required
def view_report(filename):
    try:
        # ׳‘׳“׳™׳§׳” ׳©׳”׳§׳•׳‘׳¥ ׳§׳™׳™׳
        if not os.path.exists(filename):
            flash('׳”׳“׳•׳— ׳”׳׳‘׳•׳§׳© ׳׳ ׳ ׳׳¦׳', 'danger')
            return redirect(url_for('reports'))
        
        # ׳׳”׳—׳–׳™׳¨ ׳׳× ׳”׳§׳•׳‘׳¥ ׳׳”׳•׳¨׳“׳”
        return send_file(
            filename,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'׳©׳’׳™׳׳” ׳‘׳˜׳¢׳™׳ ׳× ׳”׳“׳•׳—: {str(e)}', 'danger')
        return redirect(url_for('reports'))

# ׳™׳¦׳™׳¨׳× ׳“׳•׳— ׳׳¢׳•׳“׳›׳
@app.route('/generate_report')
@login_required
def generate_report():
    try:
        # ׳™׳¦׳™׳¨׳× ׳“׳•׳— ׳׳₪׳™ ׳×׳₪׳§׳™׳“ ׳”׳׳©׳×׳׳©
        user_role = session['user']['role']
        user_entity_id = session['user']['entity_id']
        
        # ׳™׳¦׳™׳¨׳× ׳”׳“׳•׳—
        output_file = generate_role_based_report(user_role, user_entity_id)
        
        if output_file:
            flash(f'׳”׳“׳•׳— "{output_file}" ׳ ׳•׳¦׳¨ ׳‘׳”׳¦׳׳—׳”', 'success')
        else:
            flash('׳©׳’׳™׳׳” ׳‘׳™׳¦׳™׳¨׳× ׳”׳“׳•׳— ׳”׳׳¢׳•׳“׳›׳', 'danger')
    except Exception as e:
        flash(f'׳©׳’׳™׳׳” ׳‘׳™׳¦׳™׳¨׳× ׳”׳“׳•׳—: {str(e)}', 'danger')
    
    return redirect(url_for('reports'))
# ׳ ׳™׳”׳•׳ ׳׳©׳×׳׳©׳™׳
@app.route('/users')
@admin_required
def users():
    # ׳˜׳¢׳™׳ ׳× ׳׳©׳×׳׳©׳™׳
    users_data = load_users()
    users_list = users_data['users']
    
    return render_template('users.html', 
                          users=users_list)

# ׳”׳•׳¡׳₪׳× ׳׳©׳×׳׳© ׳—׳“׳©
@app.route('/add-user', methods=['GET', 'POST'])
@admin_required
def add_user():
    excel_data = load_excel_data()
    agents = excel_data['agents']
    super_agents = excel_data['super_agents']
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        name = request.form.get('name')
        role = request.form.get('role')
        entity_id = request.form.get('entity_id') if role in ['agent', 'super_agent'] else None
        
        # ׳•׳™׳“׳•׳ ׳©׳ ׳׳©׳×׳׳© ׳™׳™׳—׳•׳“׳™
        users_data = load_users()
        if any(user['username'] == username for user in users_data['users']):
            flash('׳©׳ ׳”׳׳©׳×׳׳© ׳›׳‘׳¨ ׳§׳™׳™׳ ׳‘׳׳¢׳¨׳›׳×', 'danger')
        else:
            # ׳”׳•׳¡׳₪׳× ׳”׳׳©׳×׳׳© ׳”׳—׳“׳©
            new_user = {
                'username': username,
                'password': generate_password_hash(password),
                'name': name,
                'role': role,
                'entity_id': entity_id
            }
            
            users_data['users'].append(new_user)
            save_users(users_data)
            
            flash('׳”׳׳©׳×׳׳© ׳ ׳•׳¡׳£ ׳‘׳”׳¦׳׳—׳”', 'success')
            return redirect(url_for('users'))
    
    return render_template('add_user.html', 
                          agents=agents, 
                          super_agents=super_agents)

# ׳”׳¡׳¨׳× ׳׳©׳×׳׳©
@app.route('/delete-user/<username>', methods=['POST'])
@admin_required
def delete_user(username):
    users_data = load_users()
    
    # ׳‘׳“׳™׳§׳” ׳©׳”׳׳©׳×׳׳© ׳׳ ׳׳•׳—׳§ ׳׳× ׳¢׳¦׳׳•
    if username == session['user']['username']:
        flash('׳׳™׳ ׳ ׳™׳›׳•׳ ׳׳׳—׳•׳§ ׳׳× ׳”׳׳©׳×׳׳© ׳©׳׳', 'danger')
        return redirect(url_for('users'))
    
    # ׳”׳¡׳¨׳× ׳”׳׳©׳×׳׳©
    users_data['users'] = [user for user in users_data['users'] if user['username'] != username]
    save_users(users_data)
    
    flash('׳”׳׳©׳×׳׳© ׳”׳•׳¡׳¨ ׳‘׳”׳¦׳׳—׳”', 'success')
    return redirect(url_for('users'))

# ׳”׳’׳“׳¨׳× ׳×׳™׳§׳™׳™׳× ׳”׳×׳‘׳ ׳™׳•׳×
@app.template_filter('format_datetime')
def format_datetime(value, format='%d/%m/%Y %H:%M'):
    if not value:
        return ''
    try:
        dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        return dt.strftime(format)
    except:
        try:
            dt = datetime.strptime(value, '%Y-%m-%d')
            return dt.strftime('%d/%m/%Y')
        except:
            return value

# ׳”׳’׳“׳¨׳× ׳₪׳™׳׳˜׳¨ ׳׳₪׳•׳¨׳׳˜ ׳׳˜׳‘׳¢
@app.template_filter('format_currency')
def format_currency_filter(value):
    try:
        value = int(value)
        return f"ג‚×{value:,}"
    except (ValueError, TypeError):
        return f"ג‚×0"

# ׳₪׳•׳ ׳§׳¦׳™׳” ׳׳׳×׳—׳•׳ ׳׳¡׳“ ׳”׳ ׳×׳•׳ ׳™׳ ׳•׳”׳§׳‘׳¦׳™׳ ׳”׳“׳¨׳•׳©׳™׳
def initialize_app():
    # ׳™׳¦׳™׳¨׳× ׳׳¡׳“ ׳ ׳×׳•׳ ׳™׳ ׳׳׳©׳×׳׳©׳™׳ ׳׳ ׳׳ ׳§׳™׳™׳
    load_users()
    
    # ׳™׳¦׳™׳¨׳× ׳׳¡׳“ ׳ ׳×׳•׳ ׳™׳ ׳׳×׳©׳׳•׳׳™׳ ׳׳ ׳׳ ׳§׳™׳™׳
    if not os.path.exists(PAYMENT_HISTORY_FILE):
        save_payment_history({"payments": [], "transfers": []})
    
    # ׳‘׳“׳™׳§׳” ׳©׳§׳•׳‘׳¥ ׳”׳׳§׳¡׳ ׳§׳™׳™׳
    if not os.path.exists(EXCEL_FILE):
        print(f"׳׳–׳”׳¨׳”: ׳§׳•׳‘׳¥ ׳”׳׳§׳¡׳ {EXCEL_FILE} ׳׳ ׳ ׳׳¦׳!")

# ׳”׳₪׳¢׳׳× ׳”׳׳×׳—׳•׳ ׳‘׳¢׳׳™׳™׳× ׳”׳׳₪׳׳™׳§׳¦׳™׳”
initialize_app()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
