import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import os

# קובץ המקור
EXCEL_FILE = '/home/dyagil/payment_system/amj.xlsx'
PAYMENT_HISTORY_FILE = 'payment_history.json'

def load_payment_history():
    """טעינת היסטוריית תשלומים"""
    if os.path.exists(PAYMENT_HISTORY_FILE):
        with open(PAYMENT_HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {
            "payments": [],
            "transfers": []
        }

def calculate_hands_bonus(hands, game_type):
    """חישוב בונוס לפי כמות ידיים וסוג משחק"""
    if pd.isna(hands) or hands < 1000:
        return 0

    # חישוב כמות יחידות של 1000 ידיים (עיגול כלפי מטה)
    units = int(hands // 1000)

    # חישוב הבונוס בהתאם לסוג המשחק
    bonus_rates = {
        'NLH': 100,
        'PLO': 150,
        'PLO5': 200,
        'PLO6': 200
    }

    return units * bonus_rates.get(game_type, 0)

def generate_admin_report():
    """יצירת דוח מנהל - דוח סיכומי גביה כולל כל הנתונים"""
    try:
        print("מייצר דוח מנהל...")
        output_file = 'admin_report.xlsx'

        # קריאת נתונים מקובץ המקור
        df = pd.read_excel(EXCEL_FILE, sheet_name='game stats')

        # המרת תאריכים
        df['תאריך'] = pd.to_datetime(df['תאריך'])

        # חישוב בונוס ידיים לכל שחקן לפי סוג משחק
        hands_bonus = df.groupby(['קוד שחקן', 'שם שחקן', 'סיווג משחק'])['כמות ידיים'].sum().reset_index()
        hands_bonus['בונוס ידיים'] = hands_bonus.apply(lambda row: calculate_hands_bonus(
            row['כמות ידיים'], row['סיווג משחק']), axis=1)

        # סיכום בונוס ידיים לכל שחקן
        hands_bonus = hands_bonus.groupby(['קוד שחקן', 'שם שחקן'])['בונוס ידיים'].sum().reset_index()

        # מיזוג הנתונים חזרה לדאטהפריים המקורי
        df = pd.merge(df, hands_bonus[['קוד שחקן', 'בונוס ידיים']], on='קוד שחקן', how='left')

        # הכנת קובץ אקסל חדש
        wb = Workbook()

        # הגדרת סגנונות
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        summary_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # יצירת דף סיכומי גביה
        ws_summary = wb.active
        ws_summary.title = "סיכומי גביה"
        ws_summary.sheet_view.rightToLeft = True

        # כותרת ראשית
        ws_summary.merge_cells('A1:L1')
        title_cell = ws_summary.cell(row=1, column=1)
        title_cell.value = "דוח סיכומי גביה - מנהל מערכת"
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # טעינת היסטוריית תשלומים
        payment_history = load_payment_history()
        payments = payment_history.get('payments', [])

        # יצירת מילון נתוני תשלומים לפי מזהה שחקן
        payment_data = {}
        for payment in payments:
            player_id = str(payment['player_id'])
            if player_id not in payment_data:
                payment_data[player_id] = 0
            payment_data[player_id] += payment['amount']

        # כותרות העמודות
        headers = ['שם סופר אייג׳נט', 'שם אייג׳נט', 'קוד שחקן', 'שם שחקן', 'באלנס', 'בונוס ידיים',
                  'רייק באק', 'סה"כ לגביה', 'שולם', 'נותר לתשלום', 'רייק', 'סה"כ רייק']

        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # סיכום הקוד החסר או תיקון

        # שמירת הקובץ
        wb.save(output_file)
        print(f"דוח מנהל נוצר בהצלחה: {output_file}")
        return output_file

    except Exception as e:
        print(f"שגיאה ביצירת דוח מנהל: {str(e)}")
        return None

def generate_super_agent_report(super_agent_id):
    """יצירת דוח לסופר-אייג'נט - מציג רק נתונים השייכים לסופר-אייג'נט הספציפי"""
    try:
        # טעינת קובץ המקור
        df_players = pd.read_excel(EXCEL_FILE, sheet_name='פרוט שחקנים')
        df_stats = pd.read_excel(EXCEL_FILE, sheet_name='game stats')

        # הדפסה לצורך ניפוי באגים
        print(f"מחפש סופר-אייג'נט: {super_agent_id}")
        print(f"עמודות בדף פרוט שחקנים: {df_players.columns.tolist()[:10]}")

        # סינון הנתונים לסופר-אייג'נט הספציפי
        super_agent_players = df_players[df_players['סופר אייגנט'] == super_agent_id]
        player_ids = super_agent_players['כינוי שחקן'].tolist()

        print(f"נמצאו {len(player_ids)} שחקנים תחת סופר-אייג'נט {super_agent_id}")

        # סינון הנתונים לשחקנים של הסופר-אייג'נט
        filtered_stats = df_stats[df_stats['כינוי שחקן'].isin(player_ids)]

        # טעינת נתוני תשלומים
        payment_history = load_payment_history()
        payments = payment_history['payments']

        # סינון תשלומים רלוונטיים
        filtered_payments = [p for p in payments
                            if p.get('player_name') in player_ids or
                              p.get('super_agent_name') == super_agent_id]

        # יצירת קובץ אקסל חדש
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"super_agent_{super_agent_id}_{timestamp}.xlsx"

        with pd.ExcelWriter(output_file) as writer:
            # שמירת הנתונים המסוננים
            if not super_agent_players.empty:
                super_agent_players.to_excel(writer, sheet_name='שחקנים בניהול', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='שחקנים בניהול')

            if not filtered_stats.empty:
                filtered_stats.to_excel(writer, sheet_name='סטטיסטיקות משחק', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='סטטיסטיקות משחק')

            # יצירת גיליון תשלומים
            if filtered_payments:
                payments_df = pd.DataFrame(filtered_payments)
                payments_df.to_excel(writer, sheet_name='תשלומים', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='תשלומים')

            # הוספת גיליון סיכום
            summary_data = {
                'מדד': ['מספר שחקנים', 'סך משחקים', 'סך תשלומים'],
                'ערך': [
                    len(super_agent_players),
                    len(filtered_stats),
                    sum(p.get('amount', 0) for p in filtered_payments)
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='סיכום', index=False)

        print(f"דוח סופר-אייג'נט נוצר בהצלחה: {output_file}")
        return output_file

    except Exception as e:
        print(f"שגיאה ביצירת דוח סופר-אייג'נט: {str(e)}")
        return None

def generate_agent_report(agent_id):
    """יצירת דוח לאייג'נט - מציג רק נתונים השייכים לאייג'נט הספציפי"""
    try:
        # טעינת קובץ המקור
        df_players = pd.read_excel(EXCEL_FILE, sheet_name='פרוט שחקנים')
        df_stats = pd.read_excel(EXCEL_FILE, sheet_name='game stats')

        # הדפסה לצורך ניפוי באגים
        print(f"מחפש אייג'נט: {agent_id}")
        print(f"עמודות בדף פרוט שחקנים: {df_players.columns.tolist()[:10]}")

        # סינון הנתונים לאייג'נט הספציפי
        agent_players = df_players[df_players['אייגנט'] == agent_id]
        player_ids = agent_players['כינוי שחקן'].tolist()

        print(f"נמצאו {len(player_ids)} שחקנים תחת אייג'נט {agent_id}")

        # סינון הנתונים לשחקנים של האייג'נט
        filtered_stats = df_stats[df_stats['כינוי שחקן'].isin(player_ids)]

        # טעינת נתוני תשלומים
        payment_history = load_payment_history()
        payments = payment_history['payments']

        # סינון תשלומים רלוונטיים
        filtered_payments = [p for p in payments
                            if p.get('player_name') in player_ids or
                               p.get('agent_name') == agent_id]

        # יצירת קובץ אקסל חדש
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"agent_{agent_id}_{timestamp}.xlsx"

        with pd.ExcelWriter(output_file) as writer:
            # שמירת הנתונים המסוננים
            if not agent_players.empty:
                agent_players.to_excel(writer, sheet_name='שחקנים בניהול', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='שחקנים בניהול')

            if not filtered_stats.empty:
                filtered_stats.to_excel(writer, sheet_name='סטטיסטיקות משחק', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='סטטיסטיקות משחק')

            # יצירת גיליון תשלומים
            if filtered_payments:
                payments_df = pd.DataFrame(filtered_payments)
                payments_df.to_excel(writer, sheet_name='תשלומים', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='תשלומים')

            # הוספת גיליון סיכום
            summary_data = {
                'מדד': ['מספר שחקנים', 'סך משחקים', 'סך תשלומים'],
                'ערך': [
                    len(agent_players),
                    len(filtered_stats),
                    sum(p.get('amount', 0) for p in filtered_payments)
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='סיכום', index=False)

        print(f"דוח אייג'נט נוצר בהצלחה: {output_file}")
        return output_file

    except Exception as e:
        print(f"שגיאה ביצירת דוח אייג'נט: {str(e)}")
        return None

def generate_role_based_report(role, entity_id=None):
    """יצירת דוח לפי תפקיד המשתמש"""
    if role == 'admin':
        return generate_admin_report()
    elif role == 'super_agent':
        return generate_super_agent_report(entity_id if entity_id else "ברירת מחדל")
    elif role == 'agent':
        return generate_agent_report(entity_id if entity_id else "ברירת מחדל")
    else:
        print("תפקיד לא תקין או מזהה ישות חסר")
        return None

# הרצת הפונקציה עם דוגמאות
if __name__ == '__main__':
    # דוגמא: יצירת דוח מנהל
    generate_role_based_report('admin')

    # דוגמא: יצירת דוח סופר-אייג'נט
    # generate_role_based_report('super_agent', 'שם הסופר-אייג׳נט')

    # דוגמא: יצירת דוח אייג'נט
    # generate_role_based_report('agent', 'שם האייג׳נט')
