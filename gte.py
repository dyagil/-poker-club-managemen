# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def calculate_hands_bonus(hands, game_type):
    """חישוב בונוס לפי כמות ידיים וסוג משחק"""
    if pd.isna(hands) or hands < 1000:
        return 0
    
    # חישוב כמות יחידות של 1000 ידיים (עיגול כלפי מטה)
    units = int(hands // 1000)
    
    # חישוב הבונוס בהתאם לסוג המשחק
    bonus_rates = {
        'NLH': 100,
        'PLO': 150,
        'PLO5': 200,
        'PLO6': 200
    }
    
    return units * bonus_rates.get(game_type, 0)

def are_numbers_equal(a, b, tolerance=1e-10):
    """בודק אם שני מספרים הם הפוכים אחד לשני (a = -b) עם סובלנות לטעויות קטנות"""
    return abs(a + b) < tolerance

def create_collection_report():
    try:
        output_file = 'output.xlsx'  # הגדרת שם קובץ הפלט
        # הגדרת קידוד ברירת מחדל ל-UTF-8
        import sys
        sys.stdout.reconfigure(encoding='utf-8')
        
        # קריאת רשימת הגיליונות
        xl = pd.ExcelFile('amj.xlsx')
        print("הגיליונות בקובץ:")
        for idx, sheet in enumerate(xl.sheet_names):
            print(f"{idx}. {sheet}")
        
        # קריאת קובץ האקסל מהגיליון game stats
        sheet_name = 'game stats'
        df = pd.read_excel('amj.xlsx', sheet_name=sheet_name)
        print(f"\nקורא מידע מהגיליון: {sheet_name}")
        
        print("תוכן הקובץ:")
        print(df.head())
        
        # הדפסת כל העמודות
        print("\nשמות העמודות:")
        for idx, col in enumerate(df.columns):
            print(f"{idx}. '{col}'")
        
        # המרת תאריכים
        df['תאריך'] = pd.to_datetime(df['תאריך'])
        
        # חישוב בונוס ידיים לכל שחקן לפי סוג משחק
        hands_bonus = df.groupby(['קוד שחקן', 'שם שחקן', 'סיווג משחק'])['כמות ידיים'].sum().reset_index()
        hands_bonus['בונוס ידיים'] = hands_bonus.apply(lambda row: calculate_hands_bonus(
            row['כמות ידיים'], row['סיווג משחק']), axis=1)
        
        # סיכום בונוס ידיים לכל שחקן
        hands_bonus = hands_bonus.groupby(['קוד שחקן', 'שם שחקן'])['בונוס ידיים'].sum().reset_index()
        
        # מיזוג הנתונים חזרה לדאטהפריים המקורי
        df = pd.merge(df, hands_bonus[['קוד שחקן', 'בונוס ידיים']], 
                     on='קוד שחקן', how='left')
        
        # יצירת קובץ אקסל חדש
        wb = Workbook()
        
        # הגדרת סגנונות
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        summary_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # רשימת העמודות בסדר הנכון
        columns = [
            'קוד שחקן',
            'שם שחקן',
            'באלנס',
            'בונוס ידיים',
            'רייק באק לשחקן',
            'רייק באק סוכן',
            'מתנה',
            'יתרה מחזור קודם',
            'סה"כ לגביה',
            'שולם',
            'נותר לתשלום',
            'רייק',
            'רייק שנשאר למועדון',
            'סה"כ רייק'
        ]
        
        # קבלת רשימת הסופר אייג'נטים
        super_agents = df['שם סופר אייגנט'].unique()
        
        # מעבר על כל סופר אייג'נט
        for super_agent in super_agents:
            if pd.isna(super_agent):
                continue
                
            # יצירת גיליון חדש לסופר אייג'נט
            ws = wb.create_sheet(title=str(super_agent))
            ws.sheet_view.rightToLeft = True
            
            # סינון נתונים לסופר אייג'נט
            super_agent_data = df[df['שם סופר אייגנט'] == super_agent]
            
            # קבלת רשימת האייג'נטים
            agents = super_agent_data['שם אייגנט'].unique()
            
            current_row = 1
            agent_summaries = []
            
            # מעבר על כל אייג'נט
            for agent in agents:
                if pd.isna(agent):
                    continue
                    
                # סינון נתונים לאייג'נט הנוכחי
                agent_data = super_agent_data[super_agent_data['שם אייגנט'] == agent]
                
                # יצירת טבלת סיכום לשחקנים
                player_summary = agent_data.groupby(['קוד שחקן', 'שם שחקן']).agg({
                    'באלנס': 'sum',
                    'בונוס ידיים': 'first',
                    'סה"כ רייק באק': 'sum',
                    'גאניה לאייגנט': 'sum',
                    'רייק': 'sum'
                }).reset_index()

                # סינון שחקנים שהבאלנס והרייק שלהם שווים לאפס
                player_summary = player_summary[~((player_summary['באלנס'] == 0) & (player_summary['רייק'] == 0))]
                
                # מיון הנתונים לפי שם שחקן
                player_summary = player_summary.sort_values('שם שחקן')
                
                # הוספת עמודות נוספות
                player_summary['מתנה'] = 0
                player_summary['יתרה מחזור קודם'] = 0
                player_summary['סה"כ לגביה'] = (
                    player_summary['באלנס'] +
                    player_summary['בונוס ידיים'] +
                    player_summary['סה"כ רייק באק'] +
                    player_summary['גאניה לאייגנט'] +
                    player_summary['מתנה'] +
                    player_summary['יתרה מחזור קודם']
                )
                player_summary['שולם'] = 0
                player_summary['נותר לתשלום'] = player_summary['סה"כ לגביה']
                player_summary['רייק שנשאר למועדון'] = (
                    player_summary['רייק'] -
                    player_summary['סה"כ רייק באק'] -
                    player_summary['גאניה לאייגנט'] -
                    player_summary['מתנה']
                )
                
                # כותרת לטבלת האייג'נט
                title_cell = ws.cell(row=current_row, column=1)
                title_cell.value = f"דו״ח גביה - {agent}"
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center')
                title_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                # עיצוב שאר התאים בשורת הכותרת
                for col in range(2, 15):  # A עד O
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                current_row += 1
                
                # כותרות העמודות
                for col, header in enumerate(columns, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                current_row += 1

                # הוספת שורה עליונה עם שם האייג'נט וערכי 0
                for col, column in enumerate(columns, 1):
                    cell = ws.cell(row=current_row, column=col)
                    
                    # שדות טקסטואליים
                    if column in ['קוד שחקן', 'שם שחקן']:
                        if column == 'קוד שחקן':
                            cell.value = agent
                        else:
                            cell.value = agent
                    # שדות חישוביים
                    elif column == 'סה"כ לגביה':
                        # יצירת נוסחה לחישוב סה"כ לגביה
                        row_str = str(current_row)
                        cell.value = (
                            f'=SUM('
                            f'{get_column_letter(3)}{row_str}+'  # באלנס
                            f'{get_column_letter(4)}{row_str}+'  # בונוס ידיים
                            f'{get_column_letter(5)}{row_str}+'  # רייק באק לשחקן
                            f'{get_column_letter(6)}{row_str}+'  # רייק באק סוכן
                            f'{get_column_letter(7)}{row_str}+'  # מתנה
                            f'{get_column_letter(8)}{row_str})'  # יתרה מחזור קודם
                        )
                    elif column == 'נותר לתשלום':
                        row_str = str(current_row)
                        cell.value = f'={get_column_letter(9)}{row_str}-{get_column_letter(10)}{row_str}'  # סה"כ לגביה - שולם
                    elif column == 'רייק באק סוכן':
                        # סכום כולל של רייק באק סוכן לכל השחקנים תחת האייג'נט
                        value = agent_data['גאניה לאייגנט'].sum()
                        cell.value = value
                    # שאר השדות המספריים
                    else:
                        cell.value = 0
                    
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.number_format = '#,##0'
                
                current_row += 1
                
                # הוספת נתוני השחקנים
                for _, row in player_summary.iterrows():
                    for col, column in enumerate(columns, 1):
                        cell = ws.cell(row=current_row, column=col)
                        
                        if column == 'קוד שחקן':
                            value = row['קוד שחקן']
                        elif column == 'שם שחקן':
                            value = row['שם שחקן']
                        elif column == 'באלנס':
                            value = row['באלנס']
                            # הוספת עיצוב מותנה לבאלנס
                            if value < 0:
                                cell.font = Font(color='FF0000')  # אדום לערכים שליליים
                            elif value > 0:
                                cell.font = Font(color='008000')  # ירוק לערכים חיוביים
                        elif column == 'בונוס ידיים':
                            value = row['בונוס ידיים']
                            if value > 0:
                                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                        elif column == 'רייק באק לשחקן':
                            value = row['סה"כ רייק באק']
                            if value > 0:
                                cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                        elif column == 'רייק באק סוכן':
                            value = 0  # לא מציג רייק באק סוכן בשורות השחקנים
                        elif column == 'מתנה':
                            value = 0
                        elif column == 'יתרה מחזור קודם':
                            value = 0
                        elif column == 'סה"כ לגביה':
                            # יצירת נוסחה לחישוב סה"כ לגביה
                            row_str = str(current_row)
                            value = (
                                f'=SUM('
                                f'{get_column_letter(3)}{row_str}+'  # באלנס
                                f'{get_column_letter(4)}{row_str}+'  # בונוס ידיים
                                f'{get_column_letter(5)}{row_str}+'  # רייק באק לשחקן
                                f'{get_column_letter(6)}{row_str}+'  # רייק באק סוכן
                                f'{get_column_letter(7)}{row_str}+'  # מתנה
                                f'{get_column_letter(8)}{row_str})'  # יתרה מחזור קודם
                            )
                        elif column == 'שולם':
                            value = 0
                        elif column == 'נותר לתשלום':
                            row_str = str(current_row)
                            value = f'={get_column_letter(9)}{row_str}-{get_column_letter(10)}{row_str}'  # סה"כ לגביה - שולם
                            # הוספת עיצוב מותנה לנותר לתשלום
                            cell.font = Font(bold=True)
                        elif column == 'רייק':
                            value = row['רייק']
                            if value > 0:
                                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                        elif column == 'רייק שנשאר למועדון':
                            row_str = str(current_row)
                            value = (
                                f'={get_column_letter(12)}{row_str}-'  # רייק
                                f'{get_column_letter(5)}{row_str}-'   # רייק באק לשחקן
                                f'{get_column_letter(6)}{row_str}-'   # רייק באק סוכן
                                f'{get_column_letter(7)}{row_str}'    # מתנה
                            )
                            # הוספת עיצוב מותנה לרייק שנשאר למועדון
                            cell.font = Font(bold=True)
                        elif column == 'סה"כ רייק':
                            value = row['רייק'] if 'רייק' in row else 0
                        else:
                            cell.value = ''
                        
                        cell.value = value
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                        
                        if isinstance(value, (int, float)) or (isinstance(value, str) and value.startswith('=')):
                            cell.number_format = '#,##0'
                    
                    current_row += 1
                
                # הוספת שורת סיכום
                summary_row = current_row
                for col, column in enumerate(columns, 1):
                    cell = ws.cell(row=summary_row, column=col)
                    cell.border = border
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    
                    if column == 'שם שחקן':
                        cell.value = 'סה"כ'
                    elif column in ['באלנס', 'בונוס ידיים', 'רייק באק לשחקן', 'רייק באק סוכן', 'מתנה', 'יתרה מחזור קודם', 'סה"כ לגביה', 'שולם', 'נותר לתשלום', 'רייק שנשאר למועדון', 'רייק', 'סה"כ רייק']:  # הוספנו 'רייק'
                        # חישוב הסכום כולל את השורה הראשונה (שורת האייג'נט)
                        start_row = current_row - len(player_summary) - 1  # התחלה משורת האייג'נט
                        end_row = current_row - 1  # סיום בשורה האחרונה לפני הסיכום
                        col_letter = get_column_letter(col)
                        cell.value = f'=SUM({col_letter}{start_row}:{col_letter}{end_row})'
                    else:
                        cell.value = ''
                
                current_row += 1
                
                # הוספת שורת תשלום
                payment_row = current_row  # שורה נוכחית
                
                # הוספת קו הפרדה
                for col in range(1, 15):  # A עד O
                    separator_cell = ws.cell(row=payment_row, column=col)
                    separator_cell.border = Border(
                        top=Side(border_style='medium', color='000000'),
                        bottom=Side(border_style=None),
                        left=Side(border_style=None),
                        right=Side(border_style=None)
                    )
                
                current_row += 1  # מתקדם לשורה הבאה
                
                # עיצוב שורת התשלום
                payment_cell = ws.cell(row=current_row, column=1)
                payment_cell.value = "לתשלום:"
                payment_cell.font = Font(bold=True, size=12, color='000000')
                payment_cell.alignment = Alignment(horizontal='right', vertical='center')
                payment_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                
                # עיצוב התא השני
                second_cell = ws.cell(row=current_row, column=2)
                second_cell.value = ""  # תא ריק
                second_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                
                # הוספת סכום לתשלום (מהבאלנס)
                amount_cell = ws.cell(row=current_row, column=3)
                amount_cell.value = f'={get_column_letter(3)}{summary_row}'  # מעתיק את הבאלנס משורת הסיכום
                amount_cell.number_format = '#,##0'
                amount_cell.font = Font(bold=True, size=12)
                amount_cell.alignment = Alignment(horizontal='center', vertical='center')
                amount_cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
                
                current_row += 1  # מתקדם לשורה הבאה
                
                # הוספת שורת רייק באק סוכן
                rakeback_cell = ws.cell(row=current_row, column=1)
                rakeback_cell.value = "רייק באק סוכן:"
                rakeback_cell.font = Font(bold=True, size=12, color='000000')
                rakeback_cell.alignment = Alignment(horizontal='right', vertical='center')
                rakeback_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                
                # עיצוב התא השני
                second_cell = ws.cell(row=current_row, column=2)
                second_cell.value = ""  # תא ריק
                second_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                
                # הוספת סכום רייק באק סוכן
                rakeback_amount_cell = ws.cell(row=current_row, column=3)
                rakeback_amount_cell.value = f'={get_column_letter(6)}{summary_row}'  # מעמודת גאניה לאייגנט
                rakeback_amount_cell.number_format = '#,##0'
                rakeback_amount_cell.font = Font(bold=True, size=12)
                rakeback_amount_cell.alignment = Alignment(horizontal='center', vertical='center')
                rakeback_amount_cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

                current_row += 1

                # הוספת שורת רייק באק לשחקן
                player_rakeback_cell = ws.cell(row=current_row, column=1)
                player_rakeback_cell.value = "רייק באק לשחקן:"
                player_rakeback_cell.font = Font(bold=True, size=12, color='000000')
                player_rakeback_cell.alignment = Alignment(horizontal='right', vertical='center')
                player_rakeback_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                
                # עיצוב התא השני
                second_cell = ws.cell(row=current_row, column=2)
                second_cell.value = ""  # תא ריק
                second_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                
                # הוספת סכום רייק באק לשחקן
                player_rakeback_amount_cell = ws.cell(row=current_row, column=3)
                player_rakeback_amount_cell.value = f'={get_column_letter(5)}{summary_row}'  # מעמודת רייק באק לשחקן
                player_rakeback_amount_cell.number_format = '#,##0'
                player_rakeback_amount_cell.font = Font(bold=True, size=12)
                player_rakeback_amount_cell.alignment = Alignment(horizontal='center', vertical='center')
                player_rakeback_amount_cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

                current_row += 1

                # הוספת שורת סה"כ לגביה
                total_cell = ws.cell(row=current_row, column=1)
                total_cell.value = "סה״כ לגביה:"
                total_cell.font = Font(bold=True, size=12, color='000000')
                total_cell.alignment = Alignment(horizontal='right', vertical='center')
                total_cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                
                # עיצוב התא השני
                second_cell = ws.cell(row=current_row, column=2)
                second_cell.value = ""  # תא ריק
                second_cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                
                # הוספת סכום סה"כ לגביה (סכום של שלושת השורות הקודמות)
                total_amount_cell = ws.cell(row=current_row, column=3)
                total_amount_cell.value = f'=SUM({get_column_letter(3)}{current_row-3}:{get_column_letter(3)}{current_row-1})'  # סכום שלוש השורות הקודמות
                total_amount_cell.number_format = '#,##0'
                total_amount_cell.font = Font(bold=True, size=12)
                total_amount_cell.alignment = Alignment(horizontal='center', vertical='center')
                total_amount_cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')

                current_row += 2  # מוסיף רווח של שורה אחת
                
                # שמירת סיכום האייג'נט לטבלה המסכמת
                agent_summaries.append({
                    'שם אייגנט': agent,
                    'קוד שחקן': '',  # ריק עבור שורת סיכום
                    'שם שחקן': '',    # ריק עבור שורת סיכום
                    'באלנס': f"='{str(super_agent)}'!{get_column_letter(3)}{summary_row}",  # באלנס
                    'בונוס ידיים': f"='{str(super_agent)}'!{get_column_letter(4)}{summary_row}",  # בונוס ידיים
                    'רייק באק לשחקן': f"='{str(super_agent)}'!{get_column_letter(5)}{summary_row}",  # רייק באק לשחקן
                    'רייק באק סוכן': f"='{str(super_agent)}'!{get_column_letter(6)}{summary_row}",  # רייק באק סוכן
                    'מתנה': f"='{str(super_agent)}'!{get_column_letter(7)}{summary_row}",  # מתנה
                    'יתרה מחזור קודם': f"='{str(super_agent)}'!{get_column_letter(8)}{summary_row}",  # יתרה מחזור קודם
                    'סה"כ לגביה': f"='{str(super_agent)}'!{get_column_letter(9)}{summary_row}",  # סה"כ לגביה
                    'שולם': f"='{str(super_agent)}'!{get_column_letter(10)}{summary_row}",  # שולם
                    'נותר לתשלום': f"='{str(super_agent)}'!{get_column_letter(11)}{summary_row}",  # נותר לתשלום
                    'רייק': f"='{str(super_agent)}'!{get_column_letter(12)}{summary_row}",  # רייק
                    'רייק שנשאר למועדון': f"='{str(super_agent)}'!{get_column_letter(13)}{summary_row}",  # רייק שנשאר למועדון
                    'סה"כ רייק': f"='{str(super_agent)}'!{get_column_letter(14)}{summary_row}"  # סה"כ רייק
                })
                
            # הוספת טבלת סיכום לסופר אייג'נט
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.value = f"טבלת סיכום - {super_agent}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            title_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # עיצוב שאר התאים בשורת הכותרת
            for col in range(2, 15):  # A עד O
                cell = ws.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            current_row += 1
            
            # כותרות לטבלה המסכמת
            for col, header in enumerate(columns, 1):
                if col == 1:
                    header = 'שם אייג׳נט'
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
            summary_start_row = current_row
            
            # הוספת נתוני האייג'נטים לטבלה המסכמת
            for agent_summary in agent_summaries:
                # הוספת שם האייג'נט
                ws.cell(row=current_row, column=1).value = agent_summary['שם אייגנט']
                
                # הוספת שאר הנתונים מהעמודה השלישית (באלנס)
                for col in range(3, len(columns) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    column_name = columns[col - 1]
                    if column_name in agent_summary:
                        cell.value = agent_summary[column_name]
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                        cell.number_format = '#,##0'
                
                current_row += 1
            
            # הוספת שורת סיכום כללי
            ws.cell(row=current_row, column=1).value = 'סה"כ כללי'
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            
            # סיכום העמודות המספריות - סוכם את כל השורות בטבלה של האייג'נט
            first_row = summary_start_row
            for col in range(3, len(columns) + 1):
                cell = ws.cell(row=current_row, column=col)
                # מתחיל מהשורה הראשונה של הטבלה (אחרי הכותרות)
                start_cell = f'{get_column_letter(col)}{first_row}'
                # מסיים בשורה האחרונה של הנתונים (לפני שורת הסיכום)
                end_cell = f'{get_column_letter(col)}{current_row-1}'
                if get_column_letter(col) == 'L':  # עמודת סה"כ רייק
                    # נוסחה לחישוב סה"כ רייק - סכום של כל הרייק בטבלה
                    cell.value = f'=SUM(L{first_row}:L{current_row-1})'
                else:
                    cell.value = f'=SUM({start_cell}:{end_cell})'
                cell.font = Font(bold=True)
                cell.fill = summary_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '#,##0'
            
            current_row += 1
            
            # התאמת רוחב העמודות
            for col in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # רשימת כל הסופר אייג'נטים כולל None
        all_super_agents = df['שם סופר אייגנט'].unique()

        # מחיקת גיליונות מיותרים אם קיימים
        if 'סיכומי גביה' in wb.sheetnames:
            del wb['סיכומי גביה']
        if 'סיכומי גביה1' in wb.sheetnames:
            del wb['סיכומי גביה1']
        if 'רייק' in wb.sheetnames:
            del wb['רייק']
        if 'Sheet' in wb.sheetnames:  # מחיקת גיליון Sheet אם קיים
            del wb['Sheet']

        # יצירת לשונית סיכומי גביה
        ws_collection_summary = wb.create_sheet(title='סיכומי גביה')
        ws_collection_summary.sheet_view.rightToLeft = True

        # כותרת ראשית
        ws_collection_summary.merge_cells('A1:L1')
        title_cell = ws_collection_summary.cell(row=1, column=1)
        title_cell.value = "טבלת סיכומי גביה"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.font = Font(bold=True, color='FFFFFF', size=14)

        # כותרות העמודות
        headers = ['שם סופר אייג׳נט', 'באלנס', 'בונוס ידיים', 'רייק באק לשחקן', 'רייק באק סוכן',
                  'מתנה', 'יתרה מחזור קודם', 'סה"כ לגביה', 'שולם', 'נותר לתשלום', 'רייק', 'סה"כ רייק']

        # עיצוב כותרות
        for col, header in enumerate(headers, 1):
            cell = ws_collection_summary.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # הוספת נתונים מכל לשונית של סופר אייג'נט
        current_row = 3
        for super_agent in all_super_agents:
            if pd.isna(super_agent):
                continue

            sheet_name = str(super_agent)
            if sheet_name not in wb.sheetnames:
                continue

            ws_collection_summary.cell(row=current_row, column=1).value = super_agent

            # מציאת טבלת הסיכום בגיליון של הסופר אייג'נט
            source_sheet = wb[sheet_name]
            summary_table_start = None
            for row in range(1, source_sheet.max_row + 1):
                cell_value = source_sheet.cell(row=row, column=1).value
                if isinstance(cell_value, str) and cell_value.startswith(f"טבלת סיכום - {super_agent}"):
                    summary_table_start = row + 2  # Skip the header row
                    break

            if summary_table_start:
                # מציאת שורת סה"כ כללי בטבלת הסיכום
                total_row = None
                for row in range(summary_table_start, source_sheet.max_row + 1):
                    cell_value = source_sheet.cell(row=row, column=1).value
                    if cell_value == 'סה"כ כללי':
                        total_row = row
                        break

                if total_row:
                    # העתקת נתונים משורת הסיכום הכללי
                    for col in range(2, len(headers) + 1):
                        target_cell = ws_collection_summary.cell(row=current_row, column=col)
                        source_col = get_column_letter(col + 1)  # Adjust column mapping
                        target_cell.value = f"='{sheet_name}'!{source_col}{total_row}"
                        target_cell.number_format = '#,##0'
                        target_cell.border = border
                        target_cell.alignment = Alignment(horizontal='center')

            current_row += 1

        # הוספת שורת סיכום כללי
        summary_row = current_row
        ws_collection_summary.cell(row=summary_row, column=1).value = 'סה"כ כללי'
        ws_collection_summary.cell(row=summary_row, column=1).font = Font(bold=True)
        ws_collection_summary.cell(row=summary_row, column=1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # נוסחאות סיכום
        for col in range(2, len(headers) + 1):
            cell = ws_collection_summary.cell(row=summary_row, column=col)
            col_letter = get_column_letter(col)
            cell.value = f'=SUM({col_letter}3:{col_letter}{summary_row-1})'
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '#,##0'

        # התאמת רוחב עמודות
        for col in range(1, len(headers) + 1):
            ws_collection_summary.column_dimensions[get_column_letter(col)].width = 15

        # שמירת הקובץ
        wb.save(output_file)
        
    except Exception as e:
        print(f'שגיאה: {str(e)}')

if __name__ == '__main__':
    create_collection_report()
