# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import pandas as pd
import numpy as np
import json
import os
from datetime import datetime
import secrets
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import pytz
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from report_generator import generate_role_based_report

# הגדרת אזור זמן ישראל
IST = pytz.timezone('Asia/Jerusalem')

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # מפתח סודי לשימוש ב-session ו-flash

# קבצי נתונים
EXCEL_FILE = 'amj.xlsx'
PAYMENT_HISTORY_FILE = 'payment_history.json'
USERS_FILE = 'users.json'

# טעינת קובץ משתמשים אם קיים
def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        # יצירת משתמש מנהל ראשוני
        admin_user = {
            "username": "admin",
            "password": generate_password_hash("admin123"),
            "role": "admin",
            "name": "מנהל מערכת",
            "entity_id": None
        }
        users = {"users": [admin_user]}
        save_users(users)
        return users

def save_users(users):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, ensure_ascii=False, indent=4)

# טעינת היסטוריית תשלומים
def load_payment_history():
    if os.path.exists(PAYMENT_HISTORY_FILE):
        with open(PAYMENT_HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {
            "payments": [],
            "transfers": []
        }

def save_payment_history(history):
    with open(PAYMENT_HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=4)

# רישום תשלום חדש
def record_payment(player_id, player_name, agent_name, super_agent_name, amount, payment_date=None, method="", notes="", recorded_by=""):
    history = load_payment_history()
    
    if payment_date is None:
        payment_date = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    
    payment = {
        "player_id": player_id,
        "player_name": player_name,
        "agent_name": agent_name,
        "super_agent_name": super_agent_name,
        "amount": amount,
        "payment_date": payment_date,
        "method": method,
        "notes": notes,
        "recorded_at": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
        "recorded_by": recorded_by
    }
    
    history["payments"].append(payment)
    save_payment_history(history)
    return payment

# רישום העברת כספים
def record_transfer(from_entity, from_type, to_entity, to_type, amount, transfer_date=None, notes="", recorded_by=""):
    history = load_payment_history()
    
    if transfer_date is None:
        transfer_date = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    
    transfer = {
        "from_entity": from_entity,
        "from_type": from_type,
        "to_entity": to_entity,
        "to_type": to_type,
        "amount": amount,
        "transfer_date": transfer_date,
        "notes": notes,
        "recorded_at": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
        "recorded_by": recorded_by
    }
    
    history["transfers"].append(transfer)
    save_payment_history(history)
    return transfer

# טעינת נתונים מהאקסל
def load_excel_data():
    try:
        # קריאת גיליון game stats
        game_stats_df = pd.read_excel(EXCEL_FILE, sheet_name='game stats')
        
        # קריאת כל הגיליונות
        xl = pd.ExcelFile(EXCEL_FILE)
        sheets = xl.sheet_names
        
        # מיפוי שחקנים, אייג'נטים וסופר-אייג'נטים
        players = game_stats_df[['קוד שחקן', 'שם שחקן']].drop_duplicates().dropna().to_dict('records')
        agents = game_stats_df['שם אייגנט'].dropna().unique().tolist()
        super_agents = game_stats_df['שם סופר אייגנט'].dropna().unique().tolist()
        
        return {
            'game_stats': game_stats_df,
            'sheets': sheets,
            'players': players,
            'agents': agents,
            'super_agents': super_agents
        }
    except Exception as e:
        print(f"שגיאה בטעינת הנתונים: {str(e)}")
        return None

# חישוב סיכומים לדשבורד
def calculate_dashboard_data():
    try:
        # טעינת היסטוריית תשלומים
        history = load_payment_history()
        payments = history.get('payments', [])
        transfers = history.get('transfers', [])
        
        # טעינת נתוני Excel
        excel_data = load_excel_data()
        if not excel_data:
            return {}
            
        # חישוב סה"כ לגבייה
        try:
            # ניסיון לטעון גיליון סיכומי גבייה
            collection_summary = pd.read_excel(EXCEL_FILE, sheet_name='סיכומי גביה')
            total_to_collect = collection_summary['סה"כ לגביה'].sum()
        except:
            # אם אין גיליון כזה, מחשב מהגיליון הראשי
            game_stats = excel_data['game_stats']
            total_to_collect = game_stats['באלנס'].sum()
        
        # חישוב סה"כ ששולם
        total_paid = sum(payment['amount'] for payment in payments)
        
        # חישוב מספר שחקנים
        players_count = len(excel_data['players'])
        
        # חישוב מספר אייג'נטים
        agents_count = len(excel_data['agents'])
        
        # חישוב סה"כ העברות
        total_transfers = sum(transfer['amount'] for transfer in transfers)
        
        # תשלומים אחרונים
        last_payments = sorted(payments, key=lambda x: x['recorded_at'], reverse=True)[:5]
        
        # העברות אחרונות
        last_transfers = sorted(transfers, key=lambda x: x['recorded_at'], reverse=True)[:5]
        
        return {
            'total_to_collect': total_to_collect,
            'total_paid': total_paid,
            'players_count': players_count,
            'agents_count': agents_count,
            'total_transfers': total_transfers,
            'last_payments': last_payments,
            'last_transfers': last_transfers
        }
    except Exception as e:
        print(f"שגיאה בחישוב נתוני דשבורד: {str(e)}")
        return {}

# עדכון קובץ האקסל עם נתוני תשלומים
def update_excel_with_payments(output_file=None):
    try:
        if output_file is None:
            # יצירת קובץ פלט חדש על בסיס קובץ המקור
            filename, ext = os.path.splitext(EXCEL_FILE)
            output_file = f"{filename}_updated{ext}"
        
        # העתקת המקור לקובץ החדש
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # טעינת היסטוריית תשלומים
        history = load_payment_history()
        payments = history.get('payments', [])
        
        # עדכון עמודת 'שולם' בכל הגיליונות
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name == "Sheet" or not any(cell.value == "שולם" for row in ws.iter_rows(max_row=3) for cell in row):
                continue
                
            # מציאת העמודות הרלוונטיות
            header_row = 2  # שורת הכותרות
            paid_col = None
            player_id_col = None
            player_name_col = None
            
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=header_row, column=col).value == "שולם":
                    paid_col = col
                elif ws.cell(row=header_row, column=col).value == "קוד שחקן":
                    player_id_col = col
                elif ws.cell(row=header_row, column=col).value == "שם שחקן":
                    player_name_col = col
            
            if not (paid_col and player_id_col and player_name_col):
                continue
                
            # עדכון שורות השחקנים
            for row in range(3, ws.max_row + 1):
                player_id = ws.cell(row=row, column=player_id_col).value
                player_name = ws.cell(row=row, column=player_name_col).value
                
                if player_id and player_name:
                    # חישוב סה"כ ששולם לפי שחקן
                    total_paid = sum(
                        payment['amount'] for payment in payments 
                        if str(payment['player_id']) == str(player_id) and payment['player_name'] == player_name
                    )
                    
                    # עדכון התא
                    ws.cell(row=row, column=paid_col).value = total_paid
        
        # שמירת הקובץ
        wb.save(output_file)
        return output_file
    except Exception as e:
        print(f"שגיאה בעדכון הקובץ: {str(e)}")
        return None

# פונקציית אימות משתמש עבור צד שרת
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('נא להתחבר כדי לגשת לדף זה', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# פונקציית אימות הרשאת מנהל
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or session['user']['role'] != 'admin':
            flash('אין לך הרשאה לגשת לדף זה', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# פונקציית אימות הרשאת אייג'נט או מנהל
def agent_or_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or (session['user']['role'] != 'admin' and session['user']['role'] != 'agent' and session['user']['role'] != 'super_agent'):
            flash('אין לך הרשאה לגשת לדף זה', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

#
# ניתובי האפליקציה (Routes)
#

# דף הבית - מעבר לדשבורד אם המשתמש מחובר
@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# דף התחברות
@app.route('/login', methods=['GET', 'POST'])
def login():
    # אם המשתמש כבר מחובר, הפנה לדשבורד
    if 'user' in session:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # אימות משתמש
        users = load_users()
        for user in users['users']:
            if user['username'] == username and check_password_hash(user['password'], password):
                # שמירת פרטי המשתמש ב-session
                session['user'] = {
                    'username': user['username'],
                    'name': user['name'],
                    'role': user['role'],
                    'entity_id': user['entity_id']
                }
                
                flash(f'ברוך הבא, {user["name"]}!', 'success')
                
                # הפניה לעמוד הבא או לדשבורד
                next_page = request.args.get('next')
                if next_page:
                    return redirect(next_page)
                return redirect(url_for('dashboard'))
        
        flash('שם משתמש או סיסמה שגויים', 'danger')
    
    return render_template('login.html')

# יציאה מהמערכת
@app.route('/logout')
def logout():
    session.pop('user', None)
    flash('התנתקת מהמערכת בהצלחה', 'success')
    return redirect(url_for('login'))

# לוח מחוונים (דשבורד)
@app.route('/dashboard')
@login_required
def dashboard():
    dashboard_data = calculate_dashboard_data()
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    return render_template('dashboard.html', 
                          dashboard_data=dashboard_data, 
                          user_role=user_role,
                          user_entity_id=user_entity_id)

# רשימת שחקנים
@app.route('/players')
@login_required
def players():
    excel_data = load_excel_data()
    players_list = excel_data['players']
    
    # התאמה לפי הרשאות
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'agent':
        # סינון שחקנים השייכים לאייג'נט הנוכחי
        game_stats = excel_data['game_stats']
        agent_players = game_stats[game_stats['שם אייגנט'] == user_entity_id]['קוד שחקן'].unique()
        players_list = [p for p in players_list if p['קוד שחקן'] in agent_players]
    elif user_role == 'super_agent':
        # סינון שחקנים השייכים לסופר-אייג'נט הנוכחי
        game_stats = excel_data['game_stats']
        super_agent_players = game_stats[game_stats['שם סופר אייגנט'] == user_entity_id]['קוד שחקן'].unique()
        players_list = [p for p in players_list if p['קוד שחקן'] in super_agent_players]
    
    # טעינת נתוני תשלומים
    history = load_payment_history()
    payments = history.get('payments', [])
    
    # חישוב סה"כ ששולם לכל שחקן
    for player in players_list:
        player_id = player['קוד שחקן']
        player['total_paid'] = sum(
            payment['amount'] for payment in payments 
            if str(payment['player_id']) == str(player_id)
        )
    
    return render_template('players.html', 
                          players=players_list, 
                          user_role=user_role)

# פרטי שחקן
@app.route('/player/<player_id>')
@login_required
def player_details(player_id):
    excel_data = load_excel_data()
    game_stats = excel_data['game_stats']
    
    # מציאת נתוני השחקן
    player_data = game_stats[game_stats['קוד שחקן'].astype(str) == str(player_id)].to_dict('records')
    
    if not player_data:
        flash('שחקן לא נמצא', 'warning')
        return redirect(url_for('players'))
    
    # מידע בסיסי על השחקן
    player_info = {
        'id': player_id,
        'name': player_data[0]['שם שחקן'],
        'agent': player_data[0]['שם אייגנט'],
        'super_agent': player_data[0]['שם סופר אייגנט']
    }
    
    # טעינת היסטוריית תשלומים
    history = load_payment_history()
    
    # סינון תשלומים של השחקן
    player_payments = [
        payment for payment in history.get('payments', [])
        if str(payment['player_id']) == str(player_id)
    ]
    
    # חישוב סה"כ ששולם
    total_paid = sum(payment['amount'] for payment in player_payments)
    
    return render_template('player_details.html', 
                          player=player_info, 
                          payments=player_payments,
                          total_paid=total_paid)

# רשימת אייג'נטים
@app.route('/agents')
@login_required
def agents():
    excel_data = load_excel_data()
    agents_list = excel_data['agents']
    
    # התאמה לפי הרשאות
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'super_agent':
        # סינון אייג'נטים השייכים לסופר-אייג'נט הנוכחי
        game_stats = excel_data['game_stats']
        super_agent_agents = game_stats[game_stats['שם סופר אייגנט'] == user_entity_id]['שם אייגנט'].unique()
        agents_list = [a for a in agents_list if a in super_agent_agents]
    
    # יצירת רשימת אובייקטים עם מידע נוסף לכל אייג'נט
    agents_data = []
    game_stats = excel_data['game_stats']
    
    for agent in agents_list:
        agent_players = game_stats[game_stats['שם אייגנט'] == agent]['קוד שחקן'].unique()
        players_count = len(agent_players)
        
        # מציאת הסופר-אייג'נט
        super_agent = game_stats[game_stats['שם אייגנט'] == agent]['שם סופר אייגנט'].dropna().unique()
        super_agent_name = super_agent[0] if len(super_agent) > 0 else ""
        
        agents_data.append({
            'name': agent,
            'super_agent': super_agent_name,
            'players_count': players_count
        })
    
    return render_template('agents.html', 
                          agents=agents_data, 
                          user_role=user_role)

# פרטי אייג'נט
@app.route('/agent/<agent_name>')
@login_required
def agent_details(agent_name):
    excel_data = load_excel_data()
    game_stats = excel_data['game_stats']
    
    # מציאת שחקנים של האייג'נט
    agent_data = game_stats[game_stats['שם אייגנט'] == agent_name]
    
    if len(agent_data) == 0:
        flash('אייג\'נט לא נמצא', 'warning')
        return redirect(url_for('agents'))
    
    # רשימת שחקנים
    players = agent_data[['קוד שחקן', 'שם שחקן']].drop_duplicates().to_dict('records')
    
    # מידע על האייג'נט
    super_agent = agent_data['שם סופר אייגנט'].dropna().unique()
    agent_info = {
        'name': agent_name,
        'super_agent': super_agent[0] if len(super_agent) > 0 else "",
        'players_count': len(players)
    }
    
    # טעינת היסטוריית תשלומים
    history = load_payment_history()
    
    # סינון תשלומים של האייג'נט
    agent_payments = [
        payment for payment in history.get('payments', [])
        if payment['agent_name'] == agent_name
    ]
    
    # חישוב סה"כ ששולם
    total_paid = sum(payment['amount'] for payment in agent_payments)
    
    return render_template('agent_details.html', 
                          agent=agent_info,
                          players=players,
                          payments=agent_payments,
                          total_paid=total_paid)

# רשימת סופר-אייג'נטים
@app.route('/super-agents')
@login_required
def super_agents():
    excel_data = load_excel_data()
    super_agents_list = excel_data['super_agents']
    
    # התאמה לפי הרשאות
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'super_agent':
        # הצג רק את הסופר-אייג'נט הנוכחי
        super_agents_list = [sa for sa in super_agents_list if sa == user_entity_id]
    
    # יצירת רשימת אובייקטים עם מידע נוסף לכל סופר-אייג'נט
    super_agents_data = []
    game_stats = excel_data['game_stats']
    
    for super_agent in super_agents_list:
        agent_count = len(game_stats[game_stats['שם סופר אייגנט'] == super_agent]['שם אייגנט'].unique())
        player_count = len(game_stats[game_stats['שם סופר אייגנט'] == super_agent]['קוד שחקן'].unique())
        
        super_agents_data.append({
            'name': super_agent,
            'agents_count': agent_count,
            'players_count': player_count
        })
    
    return render_template('super_agents.html', 
                          super_agents=super_agents_data, 
                          user_role=user_role)

# פרטי סופר-אייג'נט
@app.route('/super-agent/<super_agent_name>')
@login_required
def super_agent_details(super_agent_name):
    excel_data = load_excel_data()
    game_stats = excel_data['game_stats']
    
    # מציאת נתוני הסופר-אייג'נט
    super_agent_data = game_stats[game_stats['שם סופר אייגנט'] == super_agent_name]
    
    if len(super_agent_data) == 0:
        flash('סופר-אייג\'נט לא נמצא', 'warning')
        return redirect(url_for('super_agents'))
    
    # רשימת אייג'נטים
    agents = super_agent_data['שם אייגנט'].dropna().unique().tolist()
    
    # מידע על הסופר-אייג'נט
    super_agent_info = {
        'name': super_agent_name,
        'agents_count': len(agents),
        'players_count': len(super_agent_data['קוד שחקן'].unique())
    }
    
    # טעינת היסטוריית תשלומים
    history = load_payment_history()
    
    # סינון תשלומים של הסופר-אייג'נט
    super_agent_payments = [
        payment for payment in history.get('payments', [])
        if payment['super_agent_name'] == super_agent_name
    ]
    
    # חישוב סה"כ ששולם
    total_paid = sum(payment['amount'] for payment in super_agent_payments)
    
    return render_template('super_agent_details.html', 
                          super_agent=super_agent_info,
                          agents=agents,
                          payments=super_agent_payments,
                          total_paid=total_paid)

# רישום תשלום חדש
@app.route('/add-payment', methods=['GET', 'POST'])
@agent_or_admin_required
def add_payment():
    excel_data = load_excel_data()
    players = excel_data['players']
    agents = excel_data['agents']
    super_agents = excel_data['super_agents']
    
    # התאמה לפי הרשאות
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'agent':
        # הצג רק שחקנים של האייג'נט הנוכחי
        game_stats = excel_data['game_stats']
        agent_players = game_stats[game_stats['שם אייגנט'] == user_entity_id]['קוד שחקן'].unique()
        players = [p for p in players if p['קוד שחקן'] in agent_players]
        
        # הגבל את האייג'נט להיות האייג'נט הנוכחי
        agents = [user_entity_id]
        
        # הגבל את הסופר-אייג'נט להיות הסופר-אייג'נט של האייג'נט הנוכחי
        super_agent = game_stats[game_stats['שם אייגנט'] == user_entity_id]['שם סופר אייגנט'].dropna().unique()
        super_agents = [super_agent[0]] if len(super_agent) > 0 else []
    elif user_role == 'super_agent':
        # הצג רק שחקנים של הסופר-אייג'נט הנוכחי
        game_stats = excel_data['game_stats']
        super_agent_players = game_stats[game_stats['שם סופר אייגנט'] == user_entity_id]['קוד שחקן'].unique()
        players = [p for p in players if p['קוד שחקן'] in super_agent_players]
        
        # הגבל את האייג'נט להיות האייג'נטים של הסופר-אייג'נט הנוכחי
        agents = game_stats[game_stats['שם סופר אייגנט'] == user_entity_id]['שם אייגנט'].dropna().unique().tolist()
        
        # הגבל את הסופר-אייג'נט להיות הסופר-אייג'נט הנוכחי
        super_agents = [user_entity_id]
    
    if request.method == 'POST':
        player_id = request.form.get('player_id')
        player_name = request.form.get('player_name')
        agent_name = request.form.get('agent_name')
        super_agent_name = request.form.get('super_agent_name')
        amount = float(request.form.get('amount'))
        payment_date = request.form.get('payment_date')
        method = request.form.get('method')
        notes = request.form.get('notes')
        
        # תיקון ערכים ריקים
        if not payment_date:
            payment_date = datetime.now(IST).strftime("%Y-%m-%d")
        
        # רישום התשלום
        payment = record_payment(
            player_id=player_id,
            player_name=player_name,
            agent_name=agent_name,
            super_agent_name=super_agent_name,
            amount=amount,
            payment_date=payment_date,
            method=method,
            notes=notes,
            recorded_by=session['user']['username']
        )
        
        flash('התשלום נרשם בהצלחה', 'success')
        return redirect(url_for('payments'))
    
    return render_template('add_payment.html', 
                          players=players, 
                          agents=agents, 
                          super_agents=super_agents,
                          user_role=user_role)

# רישום העברת כספים
@app.route('/add-transfer', methods=['GET', 'POST'])
@agent_or_admin_required
def add_transfer():
    excel_data = load_excel_data()
    agents = excel_data['agents']
    super_agents = excel_data['super_agents']
    
    # התאמה לפי הרשאות
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role != 'admin':
        # רק מנהל יכול לרשום העברות
        flash('אין לך הרשאה לבצע פעולה זו', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        from_type = request.form.get('from_type')
        from_entity = request.form.get('from_entity')
        to_type = request.form.get('to_type')
        to_entity = request.form.get('to_entity')
        amount = float(request.form.get('amount'))
        transfer_date = request.form.get('transfer_date')
        notes = request.form.get('notes')
        
        # תיקון ערכים ריקים
        if not transfer_date:
            transfer_date = datetime.now(IST).strftime("%Y-%m-%d")
        
        # רישום ההעברה
        transfer = record_transfer(
            from_entity=from_entity,
            from_type=from_type,
            to_entity=to_entity,
            to_type=to_type,
            amount=amount,
            transfer_date=transfer_date,
            notes=notes,
            recorded_by=session['user']['username']
        )
        
        flash('העברת הכספים נרשמה בהצלחה', 'success')
        return redirect(url_for('transfers'))
    
    return render_template('add_transfer.html', 
                          agents=agents, 
                          super_agents=super_agents)

# רשימת תשלומים
@app.route('/payments')
@login_required
def payments():
    # טעינת היסטוריית תשלומים
    history = load_payment_history()
    payments_list = history.get('payments', [])
    
    # התאמה לפי הרשאות
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    if user_role == 'agent':
        # סינון תשלומים של האייג'נט הנוכחי
        payments_list = [p for p in payments_list if p['agent_name'] == user_entity_id]
    elif user_role == 'super_agent':
        # סינון תשלומים של הסופר-אייג'נט הנוכחי
        payments_list = [p for p in payments_list if p['super_agent_name'] == user_entity_id]
    
    # מיון לפי תאריך רישום (מהחדש לישן)
    payments_list = sorted(payments_list, key=lambda x: x['recorded_at'], reverse=True)
    
    return render_template('payments.html', 
                          payments=payments_list,
                          user_role=user_role)

# רשימת העברות
@app.route('/transfers')
@login_required
def transfers():
    # טעינת היסטוריית העברות
    history = load_payment_history()
    transfers_list = history.get('transfers', [])
    
    # מיון לפי תאריך רישום (מהחדש לישן)
    transfers_list = sorted(transfers_list, key=lambda x: x['recorded_at'], reverse=True)
    
    # התאמה לפי הרשאות
    user_role = session['user']['role']
    
    if user_role != 'admin':
        # רק מנהל רואה את כל ההעברות
        transfers_list = []
    
    return render_template('transfers.html', 
                          transfers=transfers_list,
                          user_role=user_role)

# דף דוחות
@login_required
    excel_data = load_excel_data()
    sheets = excel_data['sheets']
    
    # העברת מידע על משתמש
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    return render_template('reports.html', 
                          sheets=sheets,
                          user_role=user_role,
                          user_entity_id=user_entity_id)

# תצוגת דוח ספציפי
@login_required
    try:
        # קריאת הגיליון הספציפי
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        
        # המרה לטבלת HTML
        table = df.to_html(classes='table table-striped table-bordered table-hover', index=False)
        
        return render_template('view_report.html', 
                              sheet_name=sheet_name,
                              table=table,
                              user_role=session['user']['role'])
    except Exception as e:
        flash(f'שגיאה בטעינת הדוח: {str(e)}', 'danger')
        return redirect(url_for('reports'))

# יצירת דוח מעודכן
@admin_required
    try:
        # עדכון קובץ האקסל עם נתוני התשלומים
        output_file = update_excel_with_payments()
        
        if output_file:
            flash(f'הדוח המעודכן נוצר בהצלחה: {output_file}', 'success')
        else:
            flash('שגיאה ביצירת הדוח המעודכן', 'danger')
    except Exception as e:
        flash(f'שגיאה ביצירת הדוח: {str(e)}', 'danger')
    
    return redirect(url_for('reports'))

# יצירת דוח מותאם לתפקיד
@app.route('/generate_role_report')
@login_required
def generate_role_report():
    """יצירת דוח מותאם לתפקיד המשתמש"""
    try:
        user_role = session['user']['role']
        user_entity_id = session['user']['entity_id']
        
        # יצירת הדוח
        output_file = generate_role_based_report(user_role, user_entity_id)
        
        if output_file and os.path.exists(output_file):
            # החזרת הקובץ להורדה
            return send_file(
                output_file,
                as_attachment=True,
                download_name=f"{user_role}_report.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('שגיאה ביצירת הדוח', 'danger')
            return redirect(url_for('reports'))
    
    except Exception as e:
        flash(f'שגיאה ביצירת הדוח: {str(e)}', 'danger')
        return redirect(url_for('reports'))

# דף דוחות
@app.route('/reports')
@login_required
def reports():
    # זיהוי הדוחות הקיימים
    report_files = []
    for file in os.listdir('.'):
        if file.endswith('.xlsx') and file not in [EXCEL_FILE, 'amj_updated.xlsx']:
            report_files.append({
                'filename': file,
                'display_name': file.replace('_', ' ').replace('.xlsx', ''),
                'category': 'דוח מותאם אישית',
                'modified': datetime.fromtimestamp(os.path.getmtime(file))
            })
    
    # מיון הדוחות לפי תאריך עדכון (החדש ביותר קודם)
    report_files.sort(key=lambda x: x['modified'], reverse=True)
    
    # זיהוי קטגוריות דוחות
    admin_reports = [r for r in report_files if 'admin' in r['filename'].lower()]
    super_agent_reports = [r for r in report_files if 'super_agent' in r['filename'].lower()]
    agent_reports = [r for r in report_files if 'agent_' in r['filename'].lower() and 'super' not in r['filename'].lower()]
    other_reports = [r for r in report_files if not any(x in r['filename'].lower() for x in ['admin', 'super_agent', 'agent_'])]
    
    # העברת מידע על משתמש
    user_role = session['user']['role']
    user_entity_id = session['user']['entity_id']
    
    return render_template('reports.html', 
                          admin_reports=admin_reports,
                          super_agent_reports=super_agent_reports,
                          agent_reports=agent_reports,
                          other_reports=other_reports,
                          user_role=user_role,
                          user_entity_id=user_entity_id)

# תצוגת דוח ספציפי
@app.route('/view_report/<filename>')
@login_required
def view_report(filename):
    try:
        # בדיקה שהקובץ קיים
        if not os.path.exists(filename):
            flash('הדוח המבוקש לא נמצא', 'danger')
            return redirect(url_for('reports'))
        
        # להחזיר את הקובץ להורדה
        return send_file(
            filename,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'שגיאה בטעינת הדוח: {str(e)}', 'danger')
        return redirect(url_for('reports'))

# יצירת דוח מעודכן
@app.route('/generate_report')
@login_required
def generate_report():
    try:
        # יצירת דוח לפי תפקיד המשתמש
        user_role = session['user']['role']
        user_entity_id = session['user']['entity_id']
        
        # יצירת הדוח
        output_file = generate_role_based_report(user_role, user_entity_id)
        
        if output_file:
            flash(f'הדוח "{output_file}" נוצר בהצלחה', 'success')
        else:
            flash('שגיאה ביצירת הדוח המעודכן', 'danger')
    except Exception as e:
        flash(f'שגיאה ביצירת הדוח: {str(e)}', 'danger')
    
    return redirect(url_for('reports'))
# ניהול משתמשים
@app.route('/users')
@admin_required
def users():
    # טעינת משתמשים
    users_data = load_users()
    users_list = users_data['users']
    
    return render_template('users.html', 
                          users=users_list)

# הוספת משתמש חדש
@app.route('/add-user', methods=['GET', 'POST'])
@admin_required
def add_user():
    excel_data = load_excel_data()
    agents = excel_data['agents']
    super_agents = excel_data['super_agents']
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        name = request.form.get('name')
        role = request.form.get('role')
        entity_id = request.form.get('entity_id') if role in ['agent', 'super_agent'] else None
        
        # וידוא שם משתמש ייחודי
        users_data = load_users()
        if any(user['username'] == username for user in users_data['users']):
            flash('שם המשתמש כבר קיים במערכת', 'danger')
        else:
            # הוספת המשתמש החדש
            new_user = {
                'username': username,
                'password': generate_password_hash(password),
                'name': name,
                'role': role,
                'entity_id': entity_id
            }
            
            users_data['users'].append(new_user)
            save_users(users_data)
            
            flash('המשתמש נוסף בהצלחה', 'success')
            return redirect(url_for('users'))
    
    return render_template('add_user.html', 
                          agents=agents, 
                          super_agents=super_agents)

# הסרת משתמש
@app.route('/delete-user/<username>', methods=['POST'])
@admin_required
def delete_user(username):
    users_data = load_users()
    
    # בדיקה שהמשתמש לא מוחק את עצמו
    if username == session['user']['username']:
        flash('אינך יכול למחוק את המשתמש שלך', 'danger')
        return redirect(url_for('users'))
    
    # הסרת המשתמש
    users_data['users'] = [user for user in users_data['users'] if user['username'] != username]
    save_users(users_data)
    
    flash('המשתמש הוסר בהצלחה', 'success')
    return redirect(url_for('users'))

# הגדרת תיקיית התבניות
@app.template_filter('format_datetime')
def format_datetime(value, format='%d/%m/%Y %H:%M'):
    if not value:
        return ''
    try:
        dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        return dt.strftime(format)
    except:
        try:
            dt = datetime.strptime(value, '%Y-%m-%d')
            return dt.strftime('%d/%m/%Y')
        except:
            return value

# הגדרת פילטר לפורמט מטבע
@app.template_filter('format_currency')
def format_currency_filter(value):
    try:
        value = int(value)
        return f"₪{value:,}"
    except (ValueError, TypeError):
        return f"₪0"

# פונקציה לאתחול מסד הנתונים והקבצים הדרושים
def initialize_app():
    # יצירת מסד נתונים למשתמשים אם לא קיים
    load_users()
    
    # יצירת מסד נתונים לתשלומים אם לא קיים
    if not os.path.exists(PAYMENT_HISTORY_FILE):
        save_payment_history({"payments": [], "transfers": []})
    
    # בדיקה שקובץ האקסל קיים
    if not os.path.exists(EXCEL_FILE):
        print(f"אזהרה: קובץ האקסל {EXCEL_FILE} לא נמצא!")

# הפעלת האתחול בעליית האפליקציה
initialize_app()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
